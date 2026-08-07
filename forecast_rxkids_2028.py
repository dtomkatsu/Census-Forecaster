"""RxKids Hawaiʻi — 2028 cost and benefits-by-income-quintile analysis.

Produces a single shareable workbook combining both halves of the RxKids
analysis for a hypothetical Hawaiʻi RxKids prenatal + infant cash program
under its **statutory eligibility** (qualify for the State's Medicaid
program OR family income ≤ 300% FPL for a family of applicable size
*including the expected unborn child*), projected to **2028**:

  * **Cost** — the annual fiscal outlay = the weighted sum of the expected
    RxKids benefit (total + prenatal/postnatal arms), by state and county,
    with a SDR sampling 90% CI and a parameter assumption band.
  * **Household impact** — the RxKids benefit *received* by income quintile:
    families reached, total dollars, and average benefit per family in each
    fifth of the income distribution. This is a pure distribution-of-benefits
    view; it does NOT touch the tax engine or SPM poverty thresholds, so the
    whole run is genuine TY2028 in a single pass.

The deliverable is an Excel workbook (`rxkids_2028_cost_and_impact.xlsx`)
with Summary / By-county / By-income-quintile / Assumptions / Notes tabs —
the analyst-facing format. Component CSVs and the SPM-unit parquet are
written alongside.

Method
------
1. Load PUMS (with replicate weights) and attach SPM unit IDs.
2. Project incomes forward to the target year (FPL thresholds taken at the
   same year, so income and eligibility are coherent).
3. Compute the Medicaid flag (clause 1) and the RxKids benefit (clauses
   1 + 2), aggregate to SPM units.
4. Cost = Σ rxkids_amount × weight (WGTP household weight) at SPM grain.
   Quintiles: SPM units ranked on summed income, weighted by WGTP.

Caveats
-------
* FPL for 2026-2028 is CPI-projected off the 2025 HHS table.
* Pregnancy incidence and child-age share are held at base-year values; the
  assumption band brackets this.

Usage
-----
    uv run python forecast_rxkids_2028.py \\
        --tax-year 2028 \\
        --pums-data-dir packages/data/raw/pums_2024_1yr \\
        --out reports/rxkids_2028/

    # Smoke test on the synthetic fixture (no replicate SE):
    uv run python forecast_rxkids_2028.py --use-fixture
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

LOG = logging.getLogger("forecast_rxkids_2028")
REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT / "scripts"))

PUMS_CONSTRUCTION_YEAR = 2022
WORKBOOK_NAME = "rxkids_2028_cost_and_impact.xlsx"
PDF_NAME = "rxkids_2028_cost_and_impact.pdf"
# Full Flint-design postnatal window. The base run uses the 6-month lower
# bound; this prices the optional extension to the full 12 months.
EXTENDED_POSTNATAL_MONTHS = 12
# Prenatal take-up as a fraction of postnatal (newborn) take-up. Flint
# observed ~90% prenatal enrollment vs ~98% of newborns (≈0.92). Applied so
# the prenatal arm runs below the postnatal/newborn rate; at the Flint-mature
# --takeup-rate 0.98 this recovers Flint's observed ~0.90 prenatal rate.
PRENATAL_TAKEUP_RATIO = 0.92

# ---------------------------------------------------------------------------
# Scenario set
# ---------------------------------------------------------------------------
# The forecast prices a small panel of policy designs in ONE pass, all on the
# same projected incomes and observed births and all at the run's take-up
# (default 0.90 postnatal / 0.83 prenatal). Holding take-up constant means the
# ONLY thing that moves across scenarios is the policy lever under test — the
# eligibility gate and the postnatal duration — so the cost differences are
# clean to attribute.
#
#   * statutory_*  — the bill's two-clause gate: Medicaid OR income ≤ 300% FPL
#     (income_fpl_cap=3.00); the Medicaid OR-clause still fires.
#   * universal_*  — no income/Medicaid test: income_fpl_cap=100.0 clears the
#     income clause for every family (100× FPL ≈ $2.5M for a family of 3), so
#     every birth qualifies regardless of income or Medicaid status.
#   * *_6mo / *_12mo — postnatal window: Flint's 6-month lower bound vs the full
#     12-month upper bound. Postnatal cost is linear in months; the prenatal arm
#     is unchanged ($1,500 one-time either way).
#
# $/birth at the default $1,500 prenatal + $500/mo postnatal:
#   6mo  → $1,500 + $3,000 = $4,500      12mo → $1,500 + $6,000 = $7,500
UNIVERSAL_FPL_CAP = 100.0  # ≈ no income test (every birth family qualifies)
SCENARIOS = [
    {
        "key": "statutory_6mo",
        "label": "Statutory (Medicaid OR ≤300% FPL) · 6-mo postnatal",
        "eligibility": "Medicaid OR ≤300% FPL",
        "overrides": {"income_fpl_cap": 3.00, "postnatal_months": 6},
    },
    {
        "key": "universal_6mo",
        "label": "Universal · 6-mo postnatal",
        "eligibility": "Universal (no income/Medicaid test)",
        "overrides": {"income_fpl_cap": UNIVERSAL_FPL_CAP, "postnatal_months": 6},
    },
    {
        "key": "universal_12mo",
        "label": "Universal · 12-mo postnatal",
        "eligibility": "Universal (no income/Medicaid test)",
        "overrides": {"income_fpl_cap": UNIVERSAL_FPL_CAP, "postnatal_months": 12},
    },
]
SCENARIO_BY_KEY = {s["key"]: s for s in SCENARIOS}
# The statutory 6-month design is the headline that drives the existing detailed
# Summary tab / PDF / cost_by_state.csv / cost_by_county.csv (unchanged outputs).
DEFAULT_SCENARIO_KEY = "statutory_6mo"

# Hawaiʻi resident births by year — CDC NVSR "Births: Final Data" series, by
# mother's state of residence (the same basis as the 15,535 figure cited in
# RXKIDS_METHODOLOGY.md §2 from NVSR 73-02). Used to project the birth cohort
# forward with the repo's own damped-trend ensemble, so the demographic driver
# is aged coherently with income (rather than frozen at the base-year level).
# Update with each new NVSR final-data release.
HI_BIRTHS_BY_YEAR = {
    2018: 15404,
    2019: 15403,
    2020: 15730,
    2021: 15620,
    2022: 15535,
    2023: 14643,
    2024: 14917,  # NVSR 75-02 (Jun 9 2026), Table 5, by place of residence
}
# PUMS construction / birth-observation base year (matches PUMS_CONSTRUCTION_YEAR).
BIRTH_BASE_YEAR = 2022

# ---------------------------------------------------------------------------
# Hawaiʻi DOH preliminary vital statistics — nowcast for the post-NVSR gap
# ---------------------------------------------------------------------------
# NVSR "Births: Final Data" runs ~18 months behind (2024 final published Jun
# 2026), so without a nowcast the series dead-ends at 2024 and the 2028
# projection extrapolates a 4-year gap blind. Hawaiʻi DOH (Office of Health
# Status Monitoring) publishes preliminary births MONTHLY by county at ~5 weeks'
# lag — https://health.hawaii.gov/vitalstatistics/ — which closes that gap.
#
# TWO corrections are required before DOH can join the NVSR series; neither is
# optional, and the first is NOT a fixed constant:
#
# 1. OCCURRENCE -> RESIDENCE. DOH counts births *occurring* in Hawaiʻi
#    (including to non-residents); NVSR counts births to Hawaiʻi *residents*
#    (including those occurring out of state). RxKids eligibility is a
#    residency question, so NVSR is the correct basis. The DOH/NVSR ratio is
#    strongly time-varying — it tracks travel volume, because the wedge is
#    largely non-resident births:
#        2018 1.1054 | 2019 1.0928 | 2020 1.0051 | 2021 1.0023
#        2022 1.0023 | 2023 1.0142 | 2024 1.0037
#    Pre-COVID it ran ~1.10; the travel collapse drove it to ~1.00 and it has
#    only partially recovered. We therefore calibrate on the post-COVID regime
#    ONLY (2020-2024) and carry its dispersion into the MOE rather than
#    pretending the factor is known. A pre-COVID-style recovery toward 1.10 is
#    the main directional risk and would make these nowcasts too HIGH.
# 2. TRAILING-MONTH INCOMPLETENESS. Birth certificates register with a lag, so
#    the newest month(s) in any snapshot are structurally short (in the
#    2026-07-06 pull, June reads 777 against a ~1,200 run-rate: ~65% complete).
#    Months within DOH_MATURATION_MONTHS of the snapshot are dropped, and a
#    partial year is annualised from the historical share of the retained
#    months — with the annualisation variance propagated into the MOE.
#
# Snapshot is pinned (not live-fetched) to keep runs byte-reproducible, the same
# discipline the anchor-source bundles use. Re-pull and bump DOH_SNAPSHOT when
# refreshing; drop any year that NVSR has since finalised (NVSR always wins).
DOH_SNAPSHOT = "2026-07-06"
DOH_SNAPSHOT_MONTH = 7
# Statewide births by month of OCCURRENCE, as published at DOH_SNAPSHOT.
HI_DOH_BIRTHS_MONTHLY = {
    2018: [1411, 1277, 1412, 1413, 1498, 1389, 1407, 1441, 1493, 1454, 1395, 1437],
    2019: [1391, 1342, 1353, 1337, 1449, 1338, 1448, 1480, 1443, 1417, 1431, 1403],
    2020: [1350, 1292, 1356, 1304, 1337, 1206, 1327, 1332, 1336, 1386, 1322, 1263],
    2021: [1205, 1118, 1417, 1272, 1291, 1272, 1356, 1424, 1385, 1338, 1263, 1315],
    2022: [1360, 1274, 1361, 1204, 1213, 1260, 1337, 1316, 1329, 1322, 1278, 1316],
    2023: [1321, 1167, 1241, 1197, 1213, 1235, 1218, 1325, 1222, 1284, 1217, 1211],
    2024: [1239, 1225, 1211, 1220, 1252, 1177, 1208, 1285, 1287, 1333, 1254, 1281],
    2025: [1282, 1112, 1130, 1211, 1299, 1206, 1203, 1207, 1229, 1185, 1197, 1318],
    2026: [1289, 1070, 1194, 1151, 1244, 777, 0, 0, 0, 0, 0, 0],
}
# Months nearer the snapshot than this are treated as not yet fully registered.
# 2 is empirical: in the 2026-07-06 pull May (2 months out) sits on the run-rate
# while June (1 month out) is ~35% short.
DOH_MATURATION_MONTHS = 2
# Years whose DOH/NVSR ratio defines the current (post-travel-collapse) regime.
DOH_RATIO_YEARS = (2020, 2021, 2022, 2023, 2024)

# ---------------------------------------------------------------------------
# DOH county-level births — calibration target for the county split
# ---------------------------------------------------------------------------
# The state-level birth target (NVSR + DOH nowcast, above) is well-anchored,
# but ACS PUMS's COUNTY ALLOCATION of that total is not independently anchored
# to anything and is unreliable for two distinct reasons:
#
# 1. SMALL SAMPLE. On the 2024 1-year PUMS frame, the raw (unweighted) age-0
#    infant sample is Honolulu=81, Hawaii=11, Maui=9, Kauai=6. At n=11 the
#    Poisson relative SE is ~30% -- comfortably large enough to explain
#    Hawaii County's ~36% divergence from DOH's actual (stable, ~1,900-1,950
#    every year 2018-2025) county-level count by chance alone. A larger 5-year
#    PUMS frame narrows this (n=65 for Hawaii County) but doesn't eliminate it,
#    and doesn't touch problem 2 below.
# 2. STRUCTURAL: Maui and Kauai are NOT independently sampled at all. Hawaii's
#    2020-vintage PUMA geography combines Maui + Kalawao + Kauai into a single
#    PUMA (0100) -- confirmed present, still combined, in both the 1-year and
#    5-year files (Census PUMA boundaries are fixed for the decade; no PUMS
#    vintage separates them). `PUMA0100Imputer` (analysis/puma_imputation.py)
#    assigns each unit a *probabilistic* Maui-vs-Kauai label from population
#    shares + demographic heuristics (income/family-size/housing-type), not
#    real sampled geography. Any Maui/Kauai split computed from raw PUMS
#    shares is therefore an artifact of that imputer, not survey evidence.
#
# Hawaii DOH's monthly county vital-statistics counts (same source as the
# state-level nowcast above) are genuine administrative data -- exactly
# resolved by county, immune to both problems. County shares below are the
# AGGREGATE share of each county across all complete DOH years (2018-2025;
# 2026 excluded as partial), from the same snapshot as HI_DOH_BIRTHS_MONTHLY.
# Shares are essentially flat over this window (Honolulu 72.8-74.3%, Hawaii
# 11.4-13.1%, Maui 9.35-9.79%, Kauai 4.25-5.03% year to year) -- no material
# trend, so a plain aggregate share (not recency-weighted) is the right
# estimator here, unlike the state-level birth-COUNT series, which does trend.
#
# This calibrates the SPLIT of the state BIRTH total, not the birth total
# itself -- sum(DOH_COUNTY_SHARE.values()) == 1 by construction, so summing
# the four county birth targets reproduces the state birth target exactly.
# It does NOT hold for the state COST total: eligibility is unit-specific and
# correlated with county (Hawaii County runs poorer than Honolulu), so moving
# birth mass onto a previously-undersampled county changes the state-weighted
# eligible share too (measured +2.1% on cost/recipients; see
# _calibrate_births_by_county).
DOH_COUNTY_SHARE = {
    "Honolulu": 0.733707,
    "Hawaii": 0.123490,
    "Maui": 0.095995,
    "Kauai": 0.046808,
}

# Default administrative-load fraction. Program cost in this model is pure
# benefit dollars (cash out the door); a real appropriation also carries
# operating overhead (enrollment, disbursement, eligibility, outreach).
# Cash-transfer admin loads typically run ~5-10%; reported as a separate line
# so the benefit total stays clean. Override with --admin-load.
DEFAULT_ADMIN_LOAD = 0.08

_TEAL = (31, 111, 139)
_DARK = (31, 59, 77)
_GREY = (90, 90, 90)
_LIGHT = (235, 242, 245)


def _ascii(s) -> str:
    """fpdf2 core fonts are latin-1; map the few non-latin-1 glyphs we use."""
    return (
        str(s)
        .replace("ʻ", "'").replace("–", "-").replace("—", "-")
        .replace("≤", "<=").replace("→", "->").replace("×", "x")
        .encode("latin-1", "replace").decode("latin-1")
    )

# Assumption-band sweep multipliers for the birth-rate driver. Cost is linear
# in the birth count, so the low/high multipliers bracket the band. (The
# take-up corners are computed in _assumption_band relative to the run's
# central take-up — see tk_low/tk_high there — not swept from a fixed pair.)
_SWEEP = {
    "child_under_age_share_mult": [0.75, 1.25],
}


def _parse_args(argv: Optional[list] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--tax-year", type=int, default=2028,
                   help="Target year for the analysis. Default 2028.")
    p.add_argument("--pums-data-dir", type=Path,
                   default=REPO_ROOT / "packages" / "data" / "raw" / "pums_2024_1yr",
                   help="Directory containing psam_p15 and psam_h15 PUMS files.")
    p.add_argument("--use-fixture", action="store_true", default=False,
                   help="Use the synthetic test fixture (smoke test only; "
                        "disables replicate-weight sampling CIs).")
    p.add_argument("--out", type=Path, default=REPO_ROOT / "reports" / "rxkids_2028",
                   help="Output directory.")
    p.add_argument("--no-replicate-se", action="store_true", default=False,
                   help="Skip SDR sampling standard errors even on real PUMS.")
    p.add_argument("--no-assumption-band", action="store_true", default=False,
                   help="Skip the parameter assumption-band sweep.")
    p.add_argument("--no-pdf", action="store_true", default=False,
                   help="Skip the one-page PDF summary (write only the workbook + CSVs).")
    p.add_argument("--takeup-rate", type=float, default=0.90,
                   help="Combined eligibility×claim take-up. Default 0.90 "
                        "(conservative); set 0.98 for the Flint-observed rate.")
    p.add_argument("--fertility-response", type=float, default=0.0,
                   help="Induced increase in eligible births from the cash "
                        "incentive (e.g. 0.10 for the ~10%% rise observed in "
                        "Flint). Scales both arms. Default 0.0 (static).")
    p.add_argument("--launch-operating-months", type=int, default=12,
                   help="Months the program operates in its first fiscal year "
                        "(12 = launch at FY start; e.g. 6 for a mid-year launch).")
    p.add_argument("--ramp-months", type=int, default=12,
                   help="Months for enrollment to ramp from zero to full take-up "
                        "in the launch year. Default 12 (gradual). Lower = faster.")
    p.add_argument("--admin-load", type=float, default=DEFAULT_ADMIN_LOAD,
                   help="Administrative overhead as a fraction of benefit cost "
                        "(enrollment/disbursement/eligibility). Default %.2f; "
                        "reported as a separate appropriation line." % DEFAULT_ADMIN_LOAD)
    p.add_argument("--no-birth-projection", action="store_true", default=False,
                   help="Hold the birth cohort at the base-year level instead of "
                        "projecting it to --tax-year with the damped-trend ensemble. "
                        "Still calibrates the observed infant count to vital stats.")
    p.add_argument("--no-doh-nowcast", action="store_true", default=False,
                   help="Drop the Hawaiʻi DOH preliminary-births nowcast and project "
                        "from CDC NVSR final data only. NVSR lags ~18 months, so this "
                        "leaves the post-final years unobserved (see RXKIDS_METHODOLOGY.md §3).")
    p.add_argument("--no-doh-county-shares", action="store_true", default=False,
                   help="Use raw ACS PUMS county shares for cost_by_county instead of "
                        "recalibrating to DOH's county-level vital statistics. PUMS's "
                        "county split is small-sample-noisy (Hawaii County n~11) and "
                        "Maui/Kauai are not independently sampled (combined PUMA, split "
                        "only by a demographic-heuristic imputer) — see "
                        "RXKIDS_METHODOLOGY.md §3. The state BIRTH total is unaffected "
                        "either way (redistribution, not a level change), but the state "
                        "COST total moves ~1-2% -- eligibility is unit-specific and "
                        "correlated with county, so shifting birth mass toward a "
                        "previously-undersampled (poorer) county changes the weighted "
                        "eligible share, not just cost_by_county.")
    p.add_argument("--use-proxy-births", action="store_true", default=False,
                   help="Use the legacy num_dependents × child_under_age_share birth "
                        "proxy instead of observed (age-0 dependent) infant counts. "
                        "For backward-comparison only; the proxy overstates births.")
    p.add_argument("--scenarios", type=str, default=",".join(s["key"] for s in SCENARIOS),
                   help="Comma-separated scenario keys to price. Default: all "
                        "(%s). The %r scenario always drives the detailed Summary "
                        "tab / cost_by_state.csv / cost_by_county.csv."
                        % (", ".join(s["key"] for s in SCENARIOS), DEFAULT_SCENARIO_KEY))
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args(argv)


def _selected_scenarios(arg: str) -> list:
    """Resolve the --scenarios CSV to ordered scenario specs (validated).

    Always includes DEFAULT_SCENARIO_KEY (it backs the detailed legacy outputs),
    prepended if the user omitted it. Preserves SCENARIOS order otherwise.
    """
    keys = [k.strip() for k in arg.split(",") if k.strip()]
    bad = [k for k in keys if k not in SCENARIO_BY_KEY]
    if bad:
        raise SystemExit(
            f"unknown --scenarios keys {bad}; choose from {list(SCENARIO_BY_KEY)}"
        )
    if DEFAULT_SCENARIO_KEY not in keys:
        keys = [DEFAULT_SCENARIO_KEY] + keys
    # De-dup while preserving the canonical SCENARIOS order.
    chosen = {k for k in keys}
    return [s for s in SCENARIOS if s["key"] in chosen]


# ---------------------------------------------------------------------------
# Observed births (replaces the num_dependents × rate proxy)
# ---------------------------------------------------------------------------


def _observed_births(base_units: pd.DataFrame, out_col: str = "observed_births") -> pd.DataFrame:
    """Attach a per-tax-unit OBSERVED birth count from age-0 dependents.

    PUMS does carry individual ages: the tax-unit constructor records each
    claimed dependent's real ``age`` in the ``dependents_details`` list-column.
    An age-0 dependent is an infant born within the past ~12 months — i.e. the
    annual birth flow, the correct basis for the full per-birth entitlement
    (NOT a <6-month stock, which would understate ~2×). Counting age-0
    dependents per unit (twins → 2) gives the observed births that drive both
    arms, replacing the ``num_dependents × child_under_age_share`` proxy. The
    proxy multiplies an all-dependents count (older kids, students, adult
    dependents) by a children-only-calibrated rate, overstating births
    materially.

    PERSON-WEIGHT BASIS (fixed 2026-08-06). Everything downstream multiplies a
    per-unit quantity by the unit's ``weight``, but that weight is a
    filing-status-calibrated hybrid — the household weight (WGTP) scaled by
    factors that force the tax-unit mix onto DOTAX's TY2022 filing-status shares
    (``_calculate_hybrid_weight``). That is the right basis for *revenue* and the
    wrong basis for *counting babies*: it both substitutes WGTP for the infant's
    own PWGTP and then applies a HoH/MFS/single adjustment that has nothing to do
    with demography. Weighting infants that way undercounted them ~7% and, worse,
    tilted the mix toward HoH units (factor 1.30) — i.e. toward lower-income
    single-parent families, biasing the eligible share upward.

    So we emit an **effective** count, ``Σ PWGTP(age-0 deps) / unit_weight``,
    chosen so that ``observed_births × weight`` reproduces the infants' own
    person-weighted total exactly. Every downstream consumer (cost, county rows,
    quintiles) keeps multiplying by ``weight`` unchanged and silently gets the
    right demographic basis. The raw integer head-count is preserved alongside as
    ``observed_births_n`` for reporting/QA.
    """
    if "dependents_details" not in base_units.columns:
        LOG.warning(
            "_observed_births: 'dependents_details' column absent (fixture/"
            "synthetic frame?); leaving %r unset so the proxy fallback applies.",
            out_col,
        )
        return base_units

    def _infants(details) -> list:
        if not isinstance(details, (list, tuple)):
            return []
        return [
            d for d in details
            if isinstance(d, dict) and _safe_age(d.get("age")) == 0
        ]

    def _count_infants(details) -> int:
        return len(_infants(details))

    def _infant_pwgtp(details) -> float:
        total = 0.0
        for d in _infants(details):
            try:
                v = float(d.get("pwgtp", 0.0) or 0.0)
            except (TypeError, ValueError):
                v = 0.0
            total += max(v, 0.0)
        return total

    out = base_units.copy()
    counts = out["dependents_details"].apply(_count_infants).astype(int)
    out["observed_births_n"] = counts

    if "weight" not in out.columns:
        LOG.warning("_observed_births: no 'weight' column; falling back to the "
                    "raw head-count basis (unit-weighted, not person-weighted).")
        out[out_col] = counts
        return out

    pw = out["dependents_details"].apply(_infant_pwgtp).astype(float)
    w = pd.to_numeric(out["weight"], errors="coerce").fillna(0.0).astype(float)

    if pw.sum() <= 0:
        # Frames built before 'pwgtp' was carried (or synthetic fixtures) —
        # keep the legacy head-count basis rather than silently zeroing births.
        LOG.warning("_observed_births: dependents_details carry no 'pwgtp'; "
                    "using the legacy unit-weighted head-count basis.")
        out[out_col] = counts
        return out

    eff = np.where(w > 0, pw.to_numpy(float) / np.where(w > 0, w.to_numpy(float), 1.0), 0.0)
    out[out_col] = eff
    unit_basis = float((counts * w).sum())
    LOG.info("_observed_births: %d infants; person-weighted %.0f vs "
             "legacy unit-weighted %.0f (%+.1f%%)",
             int(counts.sum()), pw.sum(), unit_basis,
             100.0 * (pw.sum() / max(unit_basis, 1.0) - 1.0))
    return out


def _safe_age(value) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return -1
    try:
        return int(value)
    except (TypeError, ValueError):
        return -1


def _doh_ratio() -> tuple:
    """Return (mean, sd) of the DOH-occurrence / NVSR-residence ratio.

    Calibrated on ``DOH_RATIO_YEARS`` only — the pre-2020 ratio (~1.10) belongs
    to a different travel regime and would bias the conversion badly. The sd is
    the honest dispersion across those years and is carried into the nowcast
    MOE; it is NOT a fixed constant we pretend to know.
    """
    import statistics

    ratios = [
        sum(HI_DOH_BIRTHS_MONTHLY[y]) / HI_BIRTHS_BY_YEAR[y]
        for y in DOH_RATIO_YEARS
        if y in HI_BIRTHS_BY_YEAR and y in HI_DOH_BIRTHS_MONTHLY
    ]
    if not ratios:
        return 1.0, 0.0
    mean = statistics.fmean(ratios)
    sd = statistics.stdev(ratios) if len(ratios) > 1 else 0.0
    return mean, sd


def _doh_mature_months(year: int) -> int:
    """How many leading months of ``year`` are old enough to be fully registered."""
    snap_year = int(DOH_SNAPSHOT[:4])
    if year < snap_year:
        return 12
    if year > snap_year:
        return 0
    return max(0, min(12, DOH_SNAPSHOT_MONTH - DOH_MATURATION_MONTHS))


def _doh_nowcast_births() -> list:
    """DOH-derived residence-basis birth nowcasts for years NVSR hasn't finalised.

    Returns a list of dicts (year, estimate, moe, mature_months, raw_occurrence,
    annualised) ready to be appended to the NVSR observation series. Each
    estimate carries three variance components so the ensemble's
    inverse-variance weighting can down-weight it honestly against a true NVSR
    final: Poisson counting noise, occurrence->residence ratio uncertainty, and
    (for partial years) annualisation uncertainty.
    """
    import math
    import statistics

    ratio, ratio_sd = _doh_ratio()
    ratio_rel_var = (ratio_sd / ratio) ** 2 if ratio else 0.0

    # Complete DOH years, used to learn the seasonal share of the first k months.
    complete = [
        y for y in sorted(HI_DOH_BIRTHS_MONTHLY)
        if _doh_mature_months(y) == 12 and sum(HI_DOH_BIRTHS_MONTHLY[y]) > 0
    ]

    out = []
    for year in sorted(HI_DOH_BIRTHS_MONTHLY):
        if year in HI_BIRTHS_BY_YEAR:
            continue  # NVSR final always wins
        k = _doh_mature_months(year)
        if k <= 0:
            continue
        months = HI_DOH_BIRTHS_MONTHLY[year][:k]
        if not months or sum(months) <= 0:
            continue
        partial = float(sum(months))

        if k == 12:
            occurrence = partial
            annual_rel_var = 0.0
        else:
            # Annualise on the historical share of the first k months.
            shares = [
                sum(HI_DOH_BIRTHS_MONTHLY[y][:k]) / sum(HI_DOH_BIRTHS_MONTHLY[y])
                for y in complete
            ]
            if not shares:
                continue
            share = statistics.fmean(shares)
            share_sd = statistics.stdev(shares) if len(shares) > 1 else 0.0
            occurrence = partial / share
            annual_rel_var = (share_sd / share) ** 2 if share else 0.0

        estimate = occurrence / ratio
        rel_var = (1.0 / estimate) + ratio_rel_var + annual_rel_var
        out.append({
            "year": year,
            "estimate": estimate,
            "moe": 1.645 * estimate * math.sqrt(rel_var),
            "mature_months": k,
            "raw_occurrence": partial,
            "annualised": k < 12,
        })
    return out


def _birth_nowcast_note(proj: dict) -> str:
    """Human-readable provenance for the DOH nowcast points, for report notes."""
    ncs = proj.get("nowcasts") or []
    if not ncs:
        return (f", from CDC NVSR finals through {proj.get('last_final_year')} only "
                f"(--no-doh-nowcast)")
    parts = []
    for n in ncs:
        tag = f"{n['year']}≈{n['estimate']:,.0f}"
        if n["annualised"]:
            tag += f" (annualised from {n['mature_months']}mo)"
        parts.append(tag)
    return (
        f". NVSR finals run through {proj['last_final_year']}; later years are "
        f"nowcast from Hawaiʻi DOH preliminary births (snapshot {proj['doh_snapshot']}) "
        f"converted occurrence→residence at ÷{proj['doh_ratio']:.4f} "
        f"(sd {proj['doh_ratio_sd']:.4f}, carried into their MOEs): "
        + "; ".join(parts)
    )


def _project_births(target_year: int, use_doh_nowcast: bool = True) -> dict:
    """Project the Hawaiʻi birth cohort to ``target_year`` via the repo ensemble.

    Feeds the CDC NVSR resident-births series (``HI_BIRTHS_BY_YEAR``) to the
    package's damped-trend + AR(1) ensemble (``project_acs_ensemble``), the same
    machinery the income forecast uses (φ=0.85/yr annual cadence). Returns the
    point projection plus its 90% PI and the base-year level. Vital-statistics
    counts are near-complete, so a small nominal Poisson SE (1.645·√n) is used
    only to let the ensemble's inverse-variance weighting run.

    When ``use_doh_nowcast`` (default), DOH preliminary counts extend the series
    past the NVSR final-data wall — see ``_doh_nowcast_births``. Those points
    carry materially wider MOEs, so the ensemble leans on them only as far as
    their precision warrants. ``--no-doh-nowcast`` restores the NVSR-only series.
    """
    import math

    from census_forecaster import AcsObservation, project_acs_ensemble

    def _obs(year, est, moe):
        return AcsObservation(
            estimate=float(est), moe=float(moe), year=year,
            vintage="1y", geoid="15", indicator="births",
        )

    obs = [
        _obs(y, v, 1.645 * math.sqrt(v))
        for y, v in sorted(HI_BIRTHS_BY_YEAR.items())
    ]
    nowcasts = _doh_nowcast_births() if use_doh_nowcast else []
    obs.extend(_obs(n["year"], n["estimate"], n["moe"]) for n in nowcasts)
    obs.sort(key=lambda o: o.year)

    base_level = float(HI_BIRTHS_BY_YEAR[BIRTH_BASE_YEAR])
    ratio, ratio_sd = _doh_ratio()
    meta = {
        "base_level": base_level,
        "nowcasts": nowcasts,
        "last_final_year": max(HI_BIRTHS_BY_YEAR),
        "doh_ratio": ratio,
        "doh_ratio_sd": ratio_sd,
        "doh_snapshot": DOH_SNAPSHOT if nowcasts else None,
    }
    fp = project_acs_ensemble(obs, target_year=target_year)
    if fp is None:
        return {"point": base_level, "ci90_low": base_level,
                "ci90_high": base_level, "projected": False, **meta}
    return {"point": float(fp.point), "ci90_low": float(fp.ci90_low),
            "ci90_high": float(fp.ci90_high), "projected": True, **meta}


def _calibrate_births_by_county(
    projected: pd.DataFrame, target: float, use_doh_county: bool = True,
) -> dict:
    """Redistribute the (already state-calibrated) birth total across counties
    to match DOH's aggregate county shares, instead of ACS PUMS's own county
    split.

    Why: the state total is well-anchored (NVSR + DOH nowcast), but PUMS's
    COUNTY allocation of it is not independently anchored to anything, for two
    distinct reasons documented at ``DOH_COUNTY_SHARE``: (1) tiny per-county
    infant samples (Hawaii County n=11 on the 2024 1yr frame — ~30% Poisson
    relative SE) and (2) Maui/Kauai are structurally NOT independently
    sampled — Hawaii's 2020 PUMA geography combines them into one PUMA, split
    only by a probabilistic demographic-heuristic imputer, not real survey
    evidence. DOH's monthly county vital-statistics are genuine administrative
    data, exactly resolved by county, immune to both problems.

    Mutates ``projected['observed_births']`` in place: rescales it per-county
    so ``Σ observed_births × weight`` within each county hits
    ``DOH_COUNTY_SHARE[county] × target``. The state BIRTH total is unchanged
    by construction (``Σ DOH_COUNTY_SHARE.values() == 1``) — this only
    reslices it. Rows in a county absent from ``DOH_COUNTY_SHARE`` keep the
    state-level scale already applied by the caller — no county is silently
    zeroed.

    NOT invariant: the state COST total. RxKids eligibility (Medicaid/FPL) is
    per-unit and correlated with county (Hawaiʻi County runs poorer than
    Honolulu), so moving birth mass from an over- to an under-sampled county
    changes how much of it lands on already-eligible vs already-ineligible
    units. Measured effect on the real frame: +2.1% on both cost and
    recipients when this recalibration is enabled, with ``eligible_families``
    (the 0/1 eligibility test itself) exactly unchanged — confirming the
    mechanism is reallocation of birth-weighted mass onto eligible units, not
    a change in who's eligible.
    """
    if (not use_doh_county or "county" not in projected.columns
            or "observed_births" not in projected.columns):
        return {"mode": "raw_pums_shares", "county_factors": {}}

    county_key = projected["county"]
    if isinstance(county_key.dtype, pd.CategoricalDtype):
        county_key = county_key.astype(object)
    county_key = county_key.fillna("Unknown")

    w = projected["weight"].fillna(0).to_numpy(dtype=float)
    b = projected["observed_births"].astype(float).to_numpy()
    scale = np.ones(len(projected), dtype=float)
    county_factors: dict = {}
    for county, share in DOH_COUNTY_SHARE.items():
        mask = (county_key == county).to_numpy()
        raw_c = float((b[mask] * w[mask]).sum())
        target_c = share * target
        factor_c = (target_c / raw_c) if raw_c > 0 else 1.0
        scale[mask] = factor_c
        county_factors[county] = {"raw": raw_c, "target": target_c, "factor": factor_c}

    unmatched = ~county_key.isin(list(DOH_COUNTY_SHARE))
    if unmatched.any():
        LOG.warning(
            "_calibrate_births_by_county: %d rows in unmapped counties (%s); "
            "left at the state-level scale, not DOH-recalibrated.",
            int(unmatched.sum()), sorted(set(county_key[unmatched])),
        )

    projected["observed_births"] = b * scale
    return {"mode": "doh_county_shares", "county_factors": county_factors}


def _calibrate_births(projected: pd.DataFrame, tax_year: int, args) -> dict:
    """Scale ``observed_births`` so the weighted total hits the (projected) target.

    Mutates ``projected['observed_births']`` in place. A single scalar factor
    folds in two corrections: (1) PUMS-infant-count → vital-statistics base
    level, and (2) the base-year → ``tax_year`` birth-cohort trend (the
    ensemble projection, unless ``--no-birth-projection``). Returns metadata
    for reporting. In ``--use-proxy-births`` mode (or when the observed column
    is absent) it is a no-op and the proxy drives both arms downstream.

    After the state-level scalar is applied, ``_calibrate_births_by_county``
    (unless ``--no-doh-county-shares``) redistributes that SAME birth total
    across counties using DOH's county shares rather than PUMS's own (noisy /
    structurally-imputed, see DOH_COUNTY_SHARE) county split. The state BIRTH
    total is unaffected either way, but the state COST total is NOT — see
    ``_calibrate_births_by_county`` for why (eligibility is unit-specific and
    correlated with county).
    """
    proj = _project_births(
        tax_year, use_doh_nowcast=not getattr(args, "no_doh_nowcast", False)
    )
    if args.use_proxy_births or "observed_births" not in projected.columns:
        return {
            "mode": "proxy", "raw_weighted": float("nan"),
            "calibrated_weighted": float("nan"), "target": float("nan"),
            "factor": 1.0, "projection": proj,
            "trend_applied": False, "county_calibration": {"mode": "n/a", "county_factors": {}},
        }

    w = projected["weight"].fillna(0).to_numpy(dtype=float)
    births = projected["observed_births"].fillna(0).to_numpy(dtype=float)
    raw_weighted = float((births * w).sum())

    trend_applied = not args.no_birth_projection and proj["projected"]
    target = proj["point"] if trend_applied else proj["base_level"]
    factor = (target / raw_weighted) if raw_weighted > 0 else 1.0
    projected["observed_births"] = projected["observed_births"].astype(float) * factor

    county_calibration = _calibrate_births_by_county(
        projected, target,
        use_doh_county=not getattr(args, "no_doh_county_shares", False),
    )

    return {
        "mode": "observed", "raw_weighted": raw_weighted,
        "calibrated_weighted": raw_weighted * factor, "target": float(target),
        "factor": factor, "projection": proj, "trend_applied": trend_applied,
        "county_calibration": county_calibration,
    }


# ---------------------------------------------------------------------------
# Cost + benefit helpers
# ---------------------------------------------------------------------------


def _weighted_cost(frame: pd.DataFrame, amount_col: str, weight_col: str = "weight") -> float:
    """Σ amount × weight over a frame (the fiscal-cost identity)."""
    if amount_col not in frame.columns:
        return 0.0
    a = frame[amount_col].fillna(0).to_numpy(dtype=float)
    w = frame[weight_col].fillna(0).to_numpy(dtype=float)
    return float((a * w).sum())


def _reached(frame: pd.DataFrame, amount_col: str, weight_col: str = "weight") -> float:
    """Weighted count of units with a positive expected benefit = the ELIGIBLE
    base (any family that clears the income/Medicaid test and has the relevant
    composition). NOT the recipient count — see _expected_recipients."""
    if amount_col not in frame.columns:
        return 0.0
    a = frame[amount_col].fillna(0).to_numpy(dtype=float)
    w = frame[weight_col].fillna(0).to_numpy(dtype=float)
    return float(w[a > 0].sum())


def _expected_recipients(
    frame: pd.DataFrame, *, pre_payment: float, post_payment: float,
    weight_col: str = "weight",
) -> tuple[float, float, float]:
    """Weighted EXPECTED recipients/year (pregnancies, infants, total).

    The arm dollar amount = (probability × per-recipient payment × take-up),
    so dividing by the full per-recipient payment recovers the expected number
    of actual claimers (probability × take-up). This is the true recipient
    count, far below the eligible base (_reached), because most eligible
    families do not have a pregnancy/infant in a given year.
    """
    w = frame[weight_col].fillna(0).to_numpy(dtype=float)
    pre = post = 0.0
    if pre_payment > 0 and "rxkids_prenatal_amount" in frame.columns:
        a = frame["rxkids_prenatal_amount"].fillna(0).to_numpy(dtype=float)
        pre = float((a / pre_payment * w).sum())
    if post_payment > 0 and "rxkids_postnatal_amount" in frame.columns:
        a = frame["rxkids_postnatal_amount"].fillna(0).to_numpy(dtype=float)
        post = float((a / post_payment * w).sum())
    return pre, post, pre + post


def _apply_fertility(frame: pd.DataFrame, fertility_response: float) -> pd.DataFrame:
    """Scale the RxKids arms by an induced-birth multiplier (1 + f).

    A behavioral fertility response (e.g. Flint's ~10% birth rise) lifts the
    number of eligible births, so it scales both the prenatal and postnatal
    arms — and therefore cost, recipients, and the quintile split — uniformly.
    """
    if not fertility_response:
        return frame
    m = 1.0 + fertility_response
    for c in ("rxkids_amount", "rxkids_prenatal_amount", "rxkids_postnatal_amount"):
        if c in frame.columns:
            frame[c] = frame[c] * m
    return frame


def _cost_with_sdr(frame: pd.DataFrame, amount_col: str) -> tuple[float, float]:
    """Return (cost, sdr_se) for an amount column on an SPM-grain frame.

    cost = Σ amount × weight (main WGTP weight); the SDR SE is computed over
    the weight_r01..weight_r80 replicate columns if present (else 0.0).
    """
    from tax_modeler.poverty.impact import (
        _detect_replicate_weight_cols,
        _sdr_se_from_replicates,
    )

    cost_0 = _weighted_cost(frame, amount_col, "weight")
    rep_cols = _detect_replicate_weight_cols(frame)
    if not rep_cols:
        return cost_0, 0.0
    a = frame[amount_col].fillna(0).to_numpy(dtype=float)
    cost_r = np.array([
        float((a * frame[rc].fillna(0).to_numpy(dtype=float)).sum())
        for rc in rep_cols
    ])
    se = float(_sdr_se_from_replicates(cost_0, cost_r))
    return cost_0, se


def _benefits_by_quintile(
    frame: pd.DataFrame, *, pre_payment: float, post_payment: float,
) -> list[dict]:
    """RxKids benefit received per weighted income quintile (SPM-unit grain).

    SPM units are ranked on summed ``income`` and split into population-equal
    fifths weighted by the WGTP household ``weight``. For each quintile:
    average income, families (weighted), expected recipients/year, total
    benefit dollars, average benefit per recipient, and share of total benefit.
    """
    from tax_modeler.metrics.distribution import weighted_ntile_labels

    if "income" not in frame.columns or frame.empty:
        return []
    labels = weighted_ntile_labels(
        frame["income"].fillna(0), frame["weight"].fillna(0), n=5, label_prefix="Q",
    ).to_numpy()
    rows = []
    for q in [f"Q{i}" for i in range(1, 6)]:
        mask = labels == q
        if not mask.any():
            continue
        sub = frame.loc[mask]
        w = sub["weight"].fillna(0).to_numpy(dtype=float)
        amt = sub["rxkids_amount"].fillna(0).to_numpy(dtype=float)
        inc = sub["income"].fillna(0).to_numpy(dtype=float)
        wsum = w.sum()
        total_benefit = float((amt * w).sum())
        _, _, recipients = _expected_recipients(
            sub, pre_payment=pre_payment, post_payment=post_payment,
        )
        rows.append({
            "quintile": q,
            "avg_income": round(float((inc * w).sum() / wsum), 0) if wsum > 0 else 0,
            "families": round(float(wsum), 0),
            "expected_recipients": round(recipients, 0),
            "total_benefit_$": round(total_benefit, 0),
            "avg_benefit_per_recipient_$": round(total_benefit / recipients, 0) if recipients > 0 else 0,
        })
    grand = sum(r["total_benefit_$"] for r in rows)
    for r in rows:
        r["share_of_benefit_pct"] = round(100.0 * r["total_benefit_$"] / grand, 1) if grand > 0 else 0.0
    return rows


def _first_year_disbursement(
    prenatal_annual: float,
    postnatal_annual: float,
    *,
    operating_months: int = 12,
    ramp_months: int = 12,
    postnatal_window: int = 6,
    start_frac: float = 0.0,
) -> dict:
    """First-fiscal-year *disbursement* (cash out the door), not benefits earned.

    A launch year differs from steady state in two ways, modeled monthly:

    * **Enrollment ramp** — take-up climbs from ``start_frac`` to full over
      ``ramp_months`` (linear). New prenatal payments scale directly with the
      enrollment level in the month they occur.
    * **Postnatal caseload fill** — each birth pays out over ``postnatal_window``
      months, so even at full enrollment the monthly postnatal caseload takes
      ``postnatal_window`` months to fill. The caseload in month t is the share
      of the trailing window already born-and-enrolled.

    Payments for late-year births that would spill past month
    ``operating_months`` fall into the next fiscal year and are excluded.
    """
    pre_m = prenatal_annual / 12.0      # steady-state monthly prenatal disbursement
    post_m = postnatal_annual / 12.0    # steady-state monthly postnatal disbursement

    def enroll(t: int) -> float:
        if ramp_months <= 0 or t >= ramp_months:
            return 1.0
        return start_frac + (1.0 - start_frac) * (t / ramp_months)

    pre_total = 0.0
    post_total = 0.0          # postnatal dollars actually paid within the FY
    post_entitlement = 0.0    # full postnatal entitlement of in-FY birth cohorts
    for t in range(1, operating_months + 1):
        pre_total += pre_m * enroll(t)
        # caseload fraction = trailing-window births that are enrolled, / window
        s = sum(enroll(t - k) for k in range(postnatal_window) if (t - k) >= 1)
        post_total += post_m * (s / postnatal_window)
        # full entitlement created by this month's birth cohort (pays out over
        # the next `postnatal_window` months; the part beyond operating_months
        # is deferred into the next fiscal year, not saved).
        post_entitlement += post_m * enroll(t)

    deferred_postnatal = max(0.0, post_entitlement - post_total)
    total = pre_total + post_total
    steady = prenatal_annual + postnatal_annual
    return {
        "prenatal": pre_total,
        "postnatal": post_total,
        "total": total,
        "deferred_postnatal": deferred_postnatal,
        "pct_of_steady": (total / steady) if steady > 0 else 0.0,
        "operating_months": operating_months,
        "ramp_months": ramp_months,
    }


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------


def _load(pir, args):
    """Load PUMS once (with replicate weights if requested) + attach SPM ids.

    Attaches the observed per-unit infant count (``observed_births``) here, on
    the base-year frame, BEFORE projection — ``dependents_details`` (the age
    source) lives on the base frame, and a plain int column rides through
    ``project_tax_units_forward`` unchanged.
    """
    want_se = not args.use_fixture and not args.no_replicate_se
    base_units, persons = pir._load_units(
        args.pums_data_dir, args.use_fixture, with_replicate_weights=want_se,
    )
    if not args.use_proxy_births:
        base_units = _observed_births(base_units)
    from tax_modeler.pipeline import enrich_with_spm_unit_id
    base_units, persons = enrich_with_spm_unit_id(base_units, persons)
    return base_units, persons


def _project(pir, base_units, year: int) -> pd.DataFrame:
    """Project base-year tax units forward to ``year`` (income aged)."""
    return pir._build_units_for_tax_year(
        base_units, year, project=(year != PUMS_CONSTRUCTION_YEAR),
    )


def _magi_proxy(units: pd.DataFrame) -> pd.Series:
    """Per-tax-unit MAGI proxy for the eligibility income test.

    Medicaid/CHIP and ACA test eligibility on Modified Adjusted Gross Income.
    The model's ``income`` is gross cash income that counts only the 85%
    taxable portion of Social Security; MAGI counts **100%**. We add back the
    non-taxable SS (``ssp_full`` − ``ssp``) for the primary and secondary
    filers. Above-the-line deductions (IRA/HSA/SE-tax/etc.) are not observable
    in PUMS, so MAGI is approximated as gross income + the SS add-back — a
    standard survey-based MAGI proxy. (Tax-exempt interest is likewise not
    separable in PUMS.)
    """
    magi = units["income"].fillna(0).astype(float).copy()
    for pre in ("primary", "secondary"):
        full = units.get(f"{pre}_ssp_full")
        taxable = units.get(f"{pre}_ssp")
        if full is not None and taxable is not None:
            magi = magi + (full.fillna(0) - taxable.fillna(0)).clip(lower=0)
    return magi


def _attach_magi_household(units: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Attach the MAGI household income + size for the FPL eligibility test.

    Medicaid/ACA define the eligibility household by tax-filing relationships
    (42 CFR 435.603) — filer + spouse + claimed tax dependents — which the
    model's TAX UNIT approximates. Because the statute anchors eligibility to
    the State's Medicaid program (clause 1), the 300% FPL test (clause 2) uses
    the same MAGI-household grain: each tax unit is tested on its own MAGI at
    its own size, NOT pooled with cohabiting partners / separate filers (the
    SPM resource-sharing family would over-pool income relative to MAGI rules).
    """
    out = units.copy()
    out["_magi"] = _magi_proxy(out)
    is_joint = (out["filing_status"] == "married_filing_jointly").astype(int)
    out["_magi_size"] = 1 + is_joint + out["num_dependents"].fillna(0).clip(lower=0).astype(int)
    return out, {"family_income_col": "_magi", "family_size_col": "_magi_size"}


def _apply_rxkids(
    units: pd.DataFrame, *, tax_year: int, overrides: Optional[dict] = None,
    birth_count_col: Optional[str] = "observed_births",
) -> pd.DataFrame:
    """Medicaid flag (clause 1) then RxKids expected benefit on a tax-unit frame.

    Eligibility (both clauses) is tested on the MAGI household ≈ the tax unit
    (filer + spouse + tax dependents), the family concept Medicaid uses —
    consistent with the statute's clause-1 Medicaid anchor.

    ``birth_count_col`` (default ``"observed_births"``) drives both arms from
    the observed/calibrated infant count when that column is present;
    ``compute_rxkids_for_units`` falls back to the proxy when it is absent
    (e.g. a ``--use-proxy-births`` run, where the column is never attached).
    """
    from tax_modeler.benefits.medicaid_hi_quest import compute_medicaid_for_units
    from tax_modeler.programs import compute_rxkids_for_units

    out, fam = _attach_magi_household(units)
    out = compute_medicaid_for_units(out, tax_year=tax_year, **fam)
    out = compute_rxkids_for_units(
        out, tax_year=tax_year, overrides=overrides,
        birth_count_col=birth_count_col, **fam,
    )
    return out


def _assumption_band(
    units, persons, *, tax_year: int, base_cost: float,
    central_takeup: float, fertility: float = 0.0, use_observed: bool = True,
    scenario_overrides: Optional[dict] = None,
) -> dict:
    """Joint (factorial-corner) assumption band over the two soft parameters.

    take-up and the birth count both scale total cost, and cost is monotone
    increasing in each, so the true min/max sit at the all-low and all-high
    corners — not a one-at-a-time perturbation. The take-up sweep brackets
    ``central_takeup`` (mostly downside, since it is capped at 1.0);
    ``fertility`` scales both corners.

    ``scenario_overrides`` (the active scenario's policy levers — e.g.
    ``income_fpl_cap``/``postnatal_months``) are merged into every corner so the
    band is computed at the SAME design as the scenario's point estimate, not the
    library defaults.

    The ±25% birth multiplier (``child_under_age_share_mult``) brackets
    birth-count uncertainty. In observed mode the birth driver is the
    ``observed_births`` column (the ``child_under_age_share`` parameter is
    inert), so the multiplier is applied to that column directly; the ±25%
    band also comfortably covers the ensemble projection's ~±23% 90% PI. In
    proxy mode it overrides ``child_under_age_share`` as before.
    """
    from tax_modeler.poverty.spm_aggregation import aggregate_to_spm_units
    from tax_modeler.programs import hawaii_rxkids_parameters

    base = hawaii_rxkids_parameters()
    tk_low = max(0.40, central_takeup * 0.80)
    tk_high = min(1.0, central_takeup * 1.02)
    cm = _SWEEP["child_under_age_share_mult"]
    fmul = 1.0 + max(0.0, fertility)

    def _corner(takeup, child_mult) -> float:
        overrides = {
            "takeup_rate": takeup,
            "prenatal_takeup_rate": round(takeup * PRENATAL_TAKEUP_RATIO, 4),
            **(scenario_overrides or {}),
        }
        u = units
        if use_observed and "observed_births" in units.columns:
            # The observed column drives births, so scale it directly.
            u = units.copy()
            u["observed_births"] = u["observed_births"].astype(float) * child_mult
        else:
            # Proxy mode: scale the rate parameter.
            overrides["child_under_age_share"] = float(
                np.clip(base.child_under_age_share * child_mult, 0.0, 1.0)
            )
        cost = _weighted_cost(
            aggregate_to_spm_units(
                _apply_rxkids(u, tax_year=tax_year, overrides=overrides), persons,
            ),
            "rxkids_amount", "weight",
        )
        return cost * fmul

    low = _corner(tk_low, min(cm))
    high = _corner(tk_high, max(cm))
    return {"min": low, "max": high, "base": base_cost,
            "takeup_range": [tk_low, tk_high], "mult_range": [min(cm), max(cm)]}


# ---------------------------------------------------------------------------
# Scenario runner
# ---------------------------------------------------------------------------


def _county_rows(frame: pd.DataFrame, *, pre_payment: float, post_payment: float) -> list[dict]:
    """Per-county cost rows for an SPM-grain frame (one row per county).

    Splits the scenario's program outlay (= dollars disbursed) and expected
    recipients across counties. Returns ``[]`` when the frame carries no county.
    """
    rows: list[dict] = []
    if "county" not in frame.columns:
        return rows
    # Fill NaN with a sentinel before grouping: groupby(dropna=False) on a key
    # with NaN builds an internal categorical grouper with a null group index,
    # which raises on some pandas versions. A sentinel avoids it everywhere.
    county_key = frame["county"]
    if isinstance(county_key.dtype, pd.CategoricalDtype):
        county_key = county_key.astype(object)
    county_key = county_key.fillna("Unknown")
    for county, grp in frame.groupby(county_key):
        c_total, c_se = _cost_with_sdr(grp, "rxkids_amount")
        _, _, grp_rec = _expected_recipients(
            grp, pre_payment=pre_payment, post_payment=post_payment,
        )
        rows.append({
            "county": county,
            "cost_total": round(c_total, 0),
            "cost_prenatal": round(_weighted_cost(grp, "rxkids_prenatal_amount"), 0),
            "cost_postnatal": round(_weighted_cost(grp, "rxkids_postnatal_amount"), 0),
            "cost_se": round(c_se, 0),
            "recipients": round(grp_rec, 0),
        })
    return rows


def _run_scenario(
    projected: pd.DataFrame, persons: pd.DataFrame, *, tax_year: int,
    scenario: dict, base_takeup: float, pre_takeup: float, fert: float,
    admin_load: float, operating_months: int, ramp_months: int,
    with_band: bool = True,
) -> dict:
    """Price one policy scenario: steady-state cost + reach + county split.

    All scenarios share the same projected incomes and observed births (passed
    in via ``projected``); they differ only by the scenario's policy
    ``overrides`` (``income_fpl_cap`` eligibility gate + ``postnatal_months``
    duration), which are merged with the run's take-up. Every per-frame metric
    reuses the existing helpers, so a scenario is just the standard pipeline run
    under a different override dict.
    """
    from tax_modeler.poverty.spm_aggregation import aggregate_to_spm_units
    from tax_modeler.programs import hawaii_rxkids_parameters

    base = hawaii_rxkids_parameters()
    ov = dict(scenario["overrides"])
    calib = {"takeup_rate": base_takeup, "prenatal_takeup_rate": pre_takeup, **ov}

    # Per-scenario payment amounts: prenatal is fixed ($1,500 one-time); the
    # postnatal per-birth entitlement scales with the scenario's month window.
    post_months = int(ov.get("postnatal_months", base.postnatal_months))
    pre_payment = base.prenatal_monthly * base.prenatal_months
    post_payment = base.postnatal_monthly_per_child * post_months

    units = _apply_rxkids(projected, tax_year=tax_year, overrides=calib)
    frame = _apply_fertility(aggregate_to_spm_units(units, persons), fert)

    total_cost, total_se = _cost_with_sdr(frame, "rxkids_amount")
    prenatal_cost, _ = _cost_with_sdr(frame, "rxkids_prenatal_amount")
    postnatal_cost, _ = _cost_with_sdr(frame, "rxkids_postnatal_amount")
    moe = 1.645 * total_se

    eligible_families = _reached(frame, "rxkids_amount")
    rec_pre, rec_inf, rec_total = _expected_recipients(
        frame, pre_payment=pre_payment, post_payment=post_payment,
    )
    avg_benefit = (total_cost / rec_total) if rec_total > 0 else 0.0

    admin_cost = total_cost * max(0.0, admin_load)
    appropriation_total = total_cost + admin_cost

    first_year = _first_year_disbursement(
        prenatal_cost, postnatal_cost, operating_months=operating_months,
        ramp_months=ramp_months, postnatal_window=post_months,
    )

    quintiles = _benefits_by_quintile(
        frame, pre_payment=pre_payment, post_payment=post_payment,
    )
    county_rows = _county_rows(frame, pre_payment=pre_payment, post_payment=post_payment)

    band = None
    if with_band:
        band = _assumption_band(
            units, persons, tax_year=tax_year, base_cost=total_cost,
            central_takeup=base_takeup, fertility=fert,
            use_observed="observed_births" in projected.columns,
            scenario_overrides=ov,
        )

    return {
        "key": scenario["key"], "label": scenario["label"],
        "eligibility": scenario["eligibility"], "postnatal_months": post_months,
        "pre_payment": pre_payment, "post_payment": post_payment,
        "per_birth": pre_payment + post_payment,
        "cost_total": total_cost, "cost_prenatal": prenatal_cost,
        "cost_postnatal": postnatal_cost, "cost_se": total_se, "moe": moe,
        "admin_cost": admin_cost, "appropriation_total": appropriation_total,
        "eligible_families": eligible_families, "rec_total": rec_total,
        "rec_pregnancies": rec_pre, "rec_infants": rec_inf,
        "avg_benefit": avg_benefit, "band": band, "first_year": first_year,
        "quintiles": quintiles, "county_rows": county_rows,
        "frame": frame, "units": units,
    }


def _scenario_summary_row(res: dict, *, tax_year: int) -> dict:
    """Flatten one scenario result into a state-level comparison CSV row."""
    band = res.get("band") or {}
    return {
        "scenario": res["key"],
        "label": res["label"],
        "tax_year": tax_year,
        "eligibility": res["eligibility"],
        "postnatal_months": res["postnatal_months"],
        "per_birth_$": round(res["per_birth"], 0),
        "cost_total_$": round(res["cost_total"], 0),
        "cost_prenatal_$": round(res["cost_prenatal"], 0),
        "cost_postnatal_$": round(res["cost_postnatal"], 0),
        "cost_se_$": round(res["cost_se"], 0),
        "cost_ci90_low_$": round(res["cost_total"] - res["moe"], 0),
        "cost_ci90_high_$": round(res["cost_total"] + res["moe"], 0),
        "assumption_band_low_$": round(band["min"], 0) if band else None,
        "assumption_band_high_$": round(band["max"], 0) if band else None,
        "admin_cost_$": round(res["admin_cost"], 0),
        "appropriation_total_$": round(res["appropriation_total"], 0),
        "eligible_families": round(res["eligible_families"], 0),
        "expected_recipients": round(res["rec_total"], 0),
        "expected_pregnancies": round(res["rec_pregnancies"], 0),
        "expected_infants": round(res["rec_infants"], 0),
        "avg_benefit_per_recipient_$": round(res["avg_benefit"], 0),
        "first_year_total_$": round(res["first_year"]["total"], 0),
    }


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------


def _write_workbook(path: Path, *, ctx: dict) -> None:
    """Assemble the analyst-facing Excel workbook from the computed context."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=14, color="1F3B4D")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F6F8B")
    label_font = Font(bold=True)
    money_fmt = "#,##0"
    pct_fmt = "0.0"
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_header_row(ws, headers):
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"RxKids Hawaiʻi — Cost & Benefits by Income Quintile, TY {ctx['tax_year']}"
    ws["A1"].font = title_font
    ws["A2"] = ("Statutory eligibility: Medicaid (clause 1) OR income ≤ 300% FPL "
                "incl. expected unborn child (clause 2)")
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A3"] = ("⚠ Point estimate — true uncertainty is ±30-40%. RxKids has no admin "
                "caseload to calibrate against; cost hinges on soft pregnancy/infant "
                "incidence assumptions. Read with the assumption band below.")
    ws["A3"].font = Font(italic=True, bold=True, color="B00000")

    rows = [
        ("FISCAL COST (annual)", None, None),
        ("Total program cost (benefit)", ctx["cost_total"], money_fmt),
        ("  Prenatal arm", ctx["cost_prenatal"], money_fmt),
        ("  Postnatal arm", ctx["cost_postnatal"], money_fmt),
        (f"  Administrative load ({100 * ctx['admin_load']:.0f}%)", ctx["admin_cost"], money_fmt),
        ("Appropriation total (benefit + admin)", ctx["appropriation_total"], money_fmt),
    ]
    if ctx["cost_se"] > 0:
        rows += [
            ("Sampling 90% CI — low", ctx["cost_total"] - ctx["moe"], money_fmt),
            ("Sampling 90% CI — high", ctx["cost_total"] + ctx["moe"], money_fmt),
            ("  SDR standard error", ctx["cost_se"], money_fmt),
        ]
    if ctx.get("band"):
        rows += [
            ("Assumption band — low", ctx["band"]["min"], money_fmt),
            ("Assumption band — high", ctx["band"]["max"], money_fmt),
        ]
    rows += [
        ("", None, None),
        (f"POTENTIAL — extend postnatal {ctx['base_postnatal_months']}→"
         f"{ctx['ext_postnatal_months']} months (optional)", None, None),
        ("  Additional cost (+6 months)", ctx["additional_cost"], money_fmt),
        ("  Total cost (12-month design)", ctx["ext_total"], money_fmt),
        ("", None, None),
        (f"FIRST FISCAL YEAR (launch) — {ctx['first_year']['operating_months']}mo "
         f"operating, {ctx['first_year']['ramp_months']}mo ramp", None, None),
        ("  Year-1 disbursement (total)", ctx["first_year"]["total"], money_fmt),
        ("    Prenatal", ctx["first_year"]["prenatal"], money_fmt),
        ("    Postnatal", ctx["first_year"]["postnatal"], money_fmt),
        ("    (postnatal deferred to next FY)", ctx["first_year"]["deferred_postnatal"], money_fmt),
        ("  Year-1 as % of steady state", round(100 * ctx["first_year"]["pct_of_steady"], 1), "0.0"),
        ("  Year-1 if ramp = 6mo (fast)", ctx["ramp_sensitivity"][6], money_fmt),
        ("  Year-1 if ramp = 18mo (slow)", ctx["ramp_sensitivity"][18], money_fmt),
        ("", None, None),
        ("HOUSEHOLD REACH", None, None),
        ("Eligible families (weighted)", ctx["eligible_families"], money_fmt),
        ("Expected recipients / year", ctx["rec_total"], money_fmt),
        ("  Expected pregnancies (prenatal)", ctx["rec_pregnancies"], money_fmt),
        ("  Expected infants (postnatal)", ctx["rec_infants"], money_fmt),
        ("Avg benefit per recipient", ctx["avg_benefit"], money_fmt),
        ("", None, None),
        ("Benefits by income quintile → see the 'By income quintile' tab.", None, None),
    ]

    r = 4
    for label, value, fmt in rows:
        cell = ws.cell(row=r, column=1, value=label)
        if value is None and fmt is None:
            cell.font = Font(bold=True, color="1F6F8B")  # section header
        else:
            cell.font = label_font if not label.startswith("  ") else Font()
            vcell = ws.cell(row=r, column=2, value=value)
            if fmt:
                vcell.number_format = fmt
        r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # ---- By income quintile ----
    wsq = wb.create_sheet("By income quintile")
    qheaders = [
        "Quintile", "Avg income ($)", "Families", "Expected recipients",
        "Total benefit ($)", "Avg benefit / recipient ($)",
        "Share of benefit (%)",
    ]
    _style_header_row(wsq, qheaders)
    for i, row in enumerate(ctx["quintiles"], start=2):
        vals = [
            row["quintile"], row["avg_income"], row["families"],
            row["expected_recipients"], row["total_benefit_$"],
            row["avg_benefit_per_recipient_$"], row["share_of_benefit_pct"],
        ]
        for j, v in enumerate(vals, start=1):
            c = wsq.cell(row=i, column=j, value=v)
            c.border = border
            if j in (2, 3, 4, 5, 6):
                c.number_format = money_fmt
            elif j == 7:
                c.number_format = pct_fmt
    for j in range(1, len(qheaders) + 1):
        wsq.column_dimensions[get_column_letter(j)].width = 22

    # ---- By county ----
    ws2 = wb.create_sheet("By county")
    headers = [
        "County", "Cost total ($)", "Cost prenatal ($)", "Cost postnatal ($)",
        "Cost SE ($)", "Expected recipients",
    ]
    _style_header_row(ws2, headers)
    for i, row in enumerate(ctx["county_rows"], start=2):
        vals = [
            row["county"], row["cost_total"], row["cost_prenatal"],
            row["cost_postnatal"], row["cost_se"], row["recipients"],
        ]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(row=i, column=j, value=v)
            c.border = border
            if j >= 2 and v is not None:
                c.number_format = money_fmt
    for j in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 18

    # ---- By scenario (state-level comparison) ----
    if ctx.get("scenario_rows"):
        wss = wb.create_sheet("By scenario")
        wss["A1"] = ("Scenario comparison — take-up held constant (0.90 newborn / "
                     "0.83 prenatal); only eligibility + postnatal months differ")
        wss["A1"].font = Font(italic=True, color="555555")
        sheaders = [
            "Scenario", "Eligibility", "Postnatal mo", "$/birth",
            "Cost total ($)", "Prenatal ($)", "Postnatal ($)",
            "Admin ($)", "Appropriation ($)", "Band low ($)", "Band high ($)",
            "Expected recipients", "Avg benefit ($)", "First-year ($)",
        ]
        for j, h in enumerate(sheaders, start=1):
            c = wss.cell(row=2, column=j, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        wss.freeze_panes = "A3"
        for i, row in enumerate(ctx["scenario_rows"], start=3):
            vals = [
                row["label"], row["eligibility"], row["postnatal_months"],
                row["per_birth_$"], row["cost_total_$"], row["cost_prenatal_$"],
                row["cost_postnatal_$"], row["admin_cost_$"],
                row["appropriation_total_$"], row["assumption_band_low_$"],
                row["assumption_band_high_$"], row["expected_recipients"],
                row["avg_benefit_per_recipient_$"], row["first_year_total_$"],
            ]
            for j, v in enumerate(vals, start=1):
                c = wss.cell(row=i, column=j, value=v)
                c.border = border
                if j >= 4 and isinstance(v, (int, float)):
                    c.number_format = money_fmt
        wss.column_dimensions["A"].width = 34
        wss.column_dimensions["B"].width = 30
        for j in range(3, len(sheaders) + 1):
            wss.column_dimensions[get_column_letter(j)].width = 16

    # ---- County × scenario (long format) ----
    if ctx.get("county_scenario_rows"):
        wsc = wb.create_sheet("County x scenario")
        cheaders = [
            "Scenario", "County", "Cost total ($)", "Cost prenatal ($)",
            "Cost postnatal ($)", "Cost SE ($)", "Expected recipients",
        ]
        _style_header_row(wsc, cheaders)
        for i, row in enumerate(ctx["county_scenario_rows"], start=2):
            vals = [
                row["label"], row["county"], row["cost_total"],
                row["cost_prenatal"], row["cost_postnatal"], row["cost_se"],
                row["recipients"],
            ]
            for j, v in enumerate(vals, start=1):
                c = wsc.cell(row=i, column=j, value=v)
                c.border = border
                if j >= 3 and v is not None:
                    c.number_format = money_fmt
        wsc.column_dimensions["A"].width = 30
        for j in range(2, len(cheaders) + 1):
            wsc.column_dimensions[get_column_letter(j)].width = 18

    # ---- Assumptions ----
    ws3 = wb.create_sheet("Assumptions")
    _style_header_row(ws3, ["Parameter", "Value", "Notes"])
    for i, (param, val, note) in enumerate(ctx["assumptions"], start=2):
        ws3.cell(row=i, column=1, value=param)
        ws3.cell(row=i, column=2, value=val)
        ws3.cell(row=i, column=3, value=note)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 60

    # ---- Notes ----
    ws4 = wb.create_sheet("Notes")
    for i, line in enumerate(ctx["notes"], start=1):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            cell.font = title_font
    ws4.column_dimensions["A"].width = 100

    wb.save(path)


# ---------------------------------------------------------------------------
# PDF (one-page shareable)
# ---------------------------------------------------------------------------


def _quintile_chart_png(quintiles: list[dict]):
    """Render a benefit-share-by-quintile bar chart to an in-memory PNG."""
    if not quintiles:
        return None
    import io

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    qs = [r["quintile"] for r in quintiles]
    shares = [r["share_of_benefit_pct"] for r in quintiles]
    fig, ax = plt.subplots(figsize=(6.4, 2.5))
    ax.bar(qs, shares, color="#1F6F8B")
    ax.set_ylabel("Share of benefit (%)")
    ax.set_xlabel("Income quintile (Q1 = lowest income)")
    for i, v in enumerate(shares):
        ax.text(i, v + 0.6, f"{v:.0f}%", ha="center", fontsize=8)
    ax.set_ylim(0, max(shares + [1]) * 1.18)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _pdf_table(pdf, *, headers, widths_frac, rows) -> None:
    """Draw a simple bordered table; first column left-aligned, rest right."""
    widths = [pdf.epw * f for f in widths_frac]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*_TEAL)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, widths):
        pdf.cell(w, 6, _ascii(h), fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    fill = False
    for row in rows:
        pdf.set_fill_color(*_LIGHT)
        for j, (val, w) in enumerate(zip(row, widths)):
            pdf.cell(w, 6, _ascii(val), border="B", fill=fill,
                     align="L" if j == 0 else "R")
        pdf.ln()
        fill = not fill


def _write_pdf(path: Path, *, ctx: dict) -> None:
    """One-page analyst-facing PDF mirroring the workbook headline."""
    from fpdf import FPDF

    def money(x):
        return f"${x:,.0f}"

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_text_color(*_DARK)
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 9, _ascii(
        f"RxKids Hawaiʻi - Cost & Benefits by Income Quintile, TY {ctx['tax_year']}"),
        new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(*_GREY)
    pdf.multi_cell(0, 5, _ascii(
        "Statutory eligibility: Medicaid (clause 1) OR income <= 300% FPL "
        "incl. expected unborn child (clause 2)"))
    pdf.set_x(pdf.l_margin)
    pdf.set_text_color(176, 0, 0)
    pdf.set_font("Helvetica", "B", 8.5)
    pdf.multi_cell(0, 4, _ascii(
        "Point estimate - true uncertainty is +/-30-40%. No admin caseload exists "
        "to calibrate against; cost hinges on soft pregnancy/infant incidence "
        "assumptions. Read with the assumption band."))
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)

    def section(title):
        pdf.set_x(pdf.l_margin)  # a preceding multi_cell can leave x at the right edge
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*_TEAL)
        pdf.cell(0, 7, _ascii(title), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

    def kv(label, value, indent=0):
        pdf.set_x(pdf.l_margin)  # always start the row at the left margin
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(95, 6, _ascii("   " * indent + label))
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, _ascii(value), new_x="LMARGIN", new_y="NEXT")

    section("Fiscal cost (annual)")
    kv("Total program cost (benefit)", money(ctx["cost_total"]))
    kv("Prenatal arm", money(ctx["cost_prenatal"]), 1)
    kv("Postnatal arm", money(ctx["cost_postnatal"]), 1)
    kv(f"Administrative load ({100 * ctx['admin_load']:.0f}%)", money(ctx["admin_cost"]), 1)
    kv("Appropriation total (benefit + admin)", money(ctx["appropriation_total"]))
    if ctx["cost_se"] > 0:
        kv("Sampling 90% CI",
           f"{money(ctx['cost_total'] - ctx['moe'])} - {money(ctx['cost_total'] + ctx['moe'])}")
    if ctx.get("band"):
        kv("Assumption band",
           f"{money(ctx['band']['min'])} - {money(ctx['band']['max'])}")
    pdf.ln(1)

    if ctx.get("scenario_rows") and len(ctx["scenario_rows"]) > 1:
        section("Scenario comparison")
        pdf.set_font("Helvetica", "I", 8.5)
        pdf.set_text_color(*_GREY)
        pdf.multi_cell(0, 4, _ascii(
            "Take-up held constant (0.90 newborn / 0.83 prenatal); only the "
            "eligibility gate and postnatal window change. Cost = annual benefit "
            "dollars; appropriation adds the admin load."))
        pdf.set_text_color(0, 0, 0)
        _pdf_table(
            pdf,
            headers=["Scenario", "$/birth", "Cost", "Appropriation", "Recipients"],
            widths_frac=[0.40, 0.13, 0.18, 0.18, 0.11],
            rows=[[
                r["label"], money(r["per_birth_$"]), money(r["cost_total_$"]),
                money(r["appropriation_total_$"]), f"{r['expected_recipients']:,.0f}",
            ] for r in ctx["scenario_rows"]],
        )
        pdf.ln(2)

    section(f"Potential option - extend postnatal {ctx['base_postnatal_months']}"
            f"-{ctx['ext_postnatal_months']} months")
    pdf.set_font("Helvetica", "I", 8.5)
    pdf.set_text_color(*_GREY)
    pdf.multi_cell(0, 4, _ascii(
        "The program may opt to extend postnatal payments to the full 12 "
        "months (Flint's upper bound). Priced here as an optional add-on:"))
    pdf.set_text_color(0, 0, 0)
    kv("Additional cost (+6 months)", money(ctx["additional_cost"]))
    kv("Total cost (12-month design)", money(ctx["ext_total"]))
    pdf.ln(1)

    fy = ctx["first_year"]
    section(f"First fiscal year (launch) - {fy['operating_months']}mo operating, "
            f"{fy['ramp_months']}mo enrollment ramp")
    pdf.set_font("Helvetica", "I", 8.5)
    pdf.set_text_color(*_GREY)
    pdf.multi_cell(0, 4, _ascii(
        "Year-1 cash disbursement, below steady state: enrollment ramps up and "
        "the 6-month postnatal caseload takes time to fill."))
    pdf.set_text_color(0, 0, 0)
    kv("Year-1 disbursement (total)", money(fy["total"]))
    kv("Prenatal / Postnatal",
       f"{money(fy['prenatal'])} / {money(fy['postnatal'])}", 1)
    kv("Postnatal deferred to next FY", money(fy["deferred_postnatal"]), 1)
    kv("Year-1 as % of steady state", f"{100 * fy['pct_of_steady']:.0f}%", 1)
    kv("Ramp sensitivity (6 / 12 / 18 mo)",
       f"{money(ctx['ramp_sensitivity'][6])} / {money(ctx['ramp_sensitivity'][12])} / "
       f"{money(ctx['ramp_sensitivity'][18])}")
    pdf.ln(1)

    section("Household reach")
    kv("Eligible families (weighted)", f"{ctx['eligible_families']:,.0f}")
    kv("Expected recipients / year", f"{ctx['rec_total']:,.0f}")
    kv("Expected pregnancies / infants",
       f"{ctx['rec_pregnancies']:,.0f} / {ctx['rec_infants']:,.0f}", 1)
    kv("Avg benefit per recipient", money(ctx["avg_benefit"]))
    pdf.ln(2)

    chart = _quintile_chart_png(ctx["quintiles"])
    if chart is not None:
        section("Benefit received by income quintile")
        pdf.image(chart, w=pdf.epw * 0.72)
        pdf.ln(2)
        _pdf_table(
            pdf,
            headers=["Quintile", "Avg income", "Families", "Recipients",
                     "Total benefit", "Share"],
            widths_frac=[0.14, 0.20, 0.17, 0.16, 0.20, 0.13],
            rows=[[
                r["quintile"], money(r["avg_income"]), f"{r['families']:,.0f}",
                f"{r['expected_recipients']:,.0f}", money(r["total_benefit_$"]),
                f"{r['share_of_benefit_pct']:.1f}%",
            ] for r in ctx["quintiles"]],
        )
        pdf.ln(3)

    if ctx["county_rows"]:
        section("Cost by county")
        _pdf_table(
            pdf,
            headers=["County", "Cost total", "Prenatal", "Postnatal", "Recipients"],
            widths_frac=[0.30, 0.18, 0.17, 0.17, 0.18],
            rows=[[
                r["county"], money(r["cost_total"]), money(r["cost_prenatal"]),
                money(r["cost_postnatal"]), f"{r['recipients']:,.0f}",
            ] for r in ctx["county_rows"]],
        )
        pdf.ln(3)

    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "I", 7.5)
    pdf.set_text_color(*_GREY)
    pdf.multi_cell(0, 4, _ascii(
        "Both arms are driven by OBSERVED births (age-0 dependents in each PUMS "
        "tax unit), calibrated to CDC NVSR resident births and projected to the "
        "target year with the repo's damped-trend ensemble. Each eligible birth "
        "draws one prenatal + one postnatal payment. Cost = the take-up-adjusted "
        "benefit weighted by the household (WGTP) weight. Eligibility is tested on "
        "MAGI at the tax-unit (MAGI-household) grain, the family Medicaid uses. "
        "Benefits-by-quintile rank SPM units on summed income into population-equal "
        "fifths. Sampling CI is ACS sampling error (SDR, 80 replicate weights); the "
        "assumption band sweeps take-up and the birth count (+/-25%). 2026-2028 FPL "
        "is CPI-projected off the 2025 HHS table. Full methodology: "
        "RXKIDS_METHODOLOGY.md."))

    pdf.output(str(path))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: Optional[list] = None) -> int:
    args = _parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    import poverty_impact_report as pir
    from tax_modeler.poverty.spm_aggregation import aggregate_to_spm_units
    from tax_modeler.programs import hawaii_rxkids_parameters

    args.out.mkdir(parents=True, exist_ok=True)
    p = hawaii_rxkids_parameters()
    ty = args.tax_year
    pre_payment = p.prenatal_monthly * p.prenatal_months
    post_payment = p.postnatal_monthly_per_child * p.postnatal_months

    LOG.info("Loading PUMS (real PUMS=%s)", not args.use_fixture)
    base_units, persons = _load(pir, args)

    LOG.info("Projecting income to TY %d", ty)
    projected = _project(pir, base_units, ty)

    # ---- Birth driver: observed infant counts, calibrated to (projected)
    #      vital statistics. Replaces the num_dependents × rate proxy, which
    #      multiplies an all-dependents count by a children-only-calibrated
    #      rate and overstates total births materially. ----
    birth_info = _calibrate_births(projected, ty, args)
    use_observed = birth_info["mode"] == "observed"

    fert = max(0.0, args.fertility_response)
    pre_takeup = round(args.takeup_rate * PRENATAL_TAKEUP_RATIO, 4)
    LOG.info("Take-up: postnatal %.2f / prenatal %.2f; fertility response %+.0f%%",
             args.takeup_rate, pre_takeup, 100 * fert)
    if use_observed:
        LOG.info("Birth driver: observed; weighted births %.0f -> calibrated %.0f "
                 "(target %.0f, ×%.3f)", birth_info["raw_weighted"],
                 birth_info["calibrated_weighted"], birth_info["target"],
                 birth_info["factor"])
    else:
        LOG.info("Birth driver: legacy proxy (num_dependents × rate)")

    # ---- Price every selected scenario on the same projected frame ----
    # Each scenario differs only by its policy overrides (eligibility gate +
    # postnatal duration); take-up and births are held constant across them, so
    # the cost differences isolate the policy levers.
    selected = _selected_scenarios(args.scenarios)
    LOG.info("Pricing %d scenario(s): %s", len(selected),
             ", ".join(s["key"] for s in selected))
    results = {}
    for sc in selected:
        LOG.info("  -> scenario %s", sc["key"])
        results[sc["key"]] = _run_scenario(
            projected, persons, tax_year=ty, scenario=sc,
            base_takeup=args.takeup_rate, pre_takeup=pre_takeup, fert=fert,
            admin_load=args.admin_load,
            operating_months=args.launch_operating_months,
            ramp_months=args.ramp_months,
            with_band=not args.no_assumption_band,
        )

    # ---- The statutory 6-month design is the headline: it drives the detailed
    #      Summary tab, PDF, cost_by_state.csv and cost_by_county.csv (unchanged
    #      outputs). Unpack it into the locals the rest of main() consumes. ----
    base_res = results[DEFAULT_SCENARIO_KEY]
    frame, units = base_res["frame"], base_res["units"]
    pre_payment, post_payment = base_res["pre_payment"], base_res["post_payment"]
    total_cost, total_se, moe = base_res["cost_total"], base_res["cost_se"], base_res["moe"]
    prenatal_cost, postnatal_cost = base_res["cost_prenatal"], base_res["cost_postnatal"]
    eligible_families = base_res["eligible_families"]
    rec_pregnancies = base_res["rec_pregnancies"]
    rec_infants, rec_total = base_res["rec_infants"], base_res["rec_total"]
    avg_benefit = base_res["avg_benefit"]
    quintiles, band = base_res["quintiles"], base_res["band"]
    county_rows = base_res["county_rows"]
    admin_cost = base_res["admin_cost"]
    appropriation_total = base_res["appropriation_total"]
    first_year = base_res["first_year"]
    LOG.info("Aggregated -> %d SPM units (baseline %s)", len(frame), DEFAULT_SCENARIO_KEY)

    # ---- Ramp sensitivity for the baseline launch year (fast / base / slow) ----
    postnatal_window = int(base_res["postnatal_months"])
    ramp_sensitivity = {
        rm: _first_year_disbursement(
            prenatal_cost, postnatal_cost,
            operating_months=args.launch_operating_months,
            ramp_months=rm, postnatal_window=postnatal_window,
        )["total"]
        for rm in (6, 12, 18)
    }

    # ---- Potential option: extend the STATUTORY design 6 -> 12 months, priced
    #      as an optional add-on (postnatal is linear in months). This is the
    #      statutory-eligibility extension; universal_12mo is the universal one. ----
    ext_overrides = {
        "takeup_rate": args.takeup_rate, "prenatal_takeup_rate": pre_takeup,
        **SCENARIO_BY_KEY[DEFAULT_SCENARIO_KEY]["overrides"],
        "postnatal_months": EXTENDED_POSTNATAL_MONTHS,
    }
    ext_units = _apply_rxkids(projected, tax_year=ty, overrides=ext_overrides)
    ext_frame = _apply_fertility(aggregate_to_spm_units(ext_units, persons), fert)
    ext_total = _weighted_cost(ext_frame, "rxkids_amount")
    ext_postnatal = _weighted_cost(ext_frame, "rxkids_postnatal_amount")
    additional_cost = ext_total - total_cost

    # ---- Scenario comparison rows (state) + county × scenario (long format) ----
    scenario_rows = [_scenario_summary_row(results[s["key"]], tax_year=ty) for s in selected]
    county_scenario_rows = [
        {"scenario": s["key"], "label": s["label"], **cr}
        for s in selected for cr in results[s["key"]]["county_rows"]
    ]

    # ---- Assemble context ----
    assumptions = [
        ("Eligibility cap (income_fpl_cap)", f"{p.income_fpl_cap:.2f}× FPL", "Clause 2: 300% FPL income test"),
        ("Medicaid OR-clause", "on", "Clause 1: medicaid_receives from compute_medicaid_for_units"),
        ("Prenatal payment", f"${p.prenatal_monthly:,.0f} × {p.prenatal_months}", "One-time per pregnancy"),
        ("Postnatal payment", f"${p.postnatal_monthly_per_child:,.0f}/mo × {p.postnatal_months}", "Per infant under cutoff"),
        ("Take-up — postnatal (newborn)", args.takeup_rate,
         "Flint observed 0.98; swept for the band"),
        ("Take-up — prenatal", pre_takeup,
         f"{PRENATAL_TAKEUP_RATIO:.2f}× postnatal (Flint: ~90% prenatal vs 98% newborn)"),
        ("Fertility response", f"{100 * fert:+.0f}%",
         "Induced eligible-birth increase (Flint ~+10%); scales both arms"),
        ("Birth driver", "observed (age-0 deps)" if use_observed else "proxy (n_dep × rate)",
         "Observed: age-0 dependents in PUMS, calibrated to NVSR; drives both arms; swept ±25%"
         if use_observed else "Legacy proxy: num_dependents × child_under_age_share"),
        ("Birth cohort — base / target",
         f"{birth_info['raw_weighted']:,.0f} -> {birth_info['target']:,.0f}"
         if use_observed else "n/a",
         (f"NVSR {BIRTH_BASE_YEAR} basis; "
          + (f"ensemble-projected to {ty}" if birth_info.get("trend_applied")
             else "held at base level")) if use_observed else "proxy mode"),
        ("Administrative load", f"{100 * args.admin_load:.0f}%",
         "Operating overhead on benefit cost; separate appropriation line"),
        ("Pregnancy Medicaid pathway", f"{p.pregnant_fpl_cap:.2f}× FPL",
         "Clause 1 prenatal pathway (196% FPL); subsumed by the 300% income test"),
        ("FPL table year", ty, "benefits/_fpl.py (2025 published; 2026-28 CPI-projected)"),
        ("Tax treatment", "non-taxable", "Added to SPM resources only; no AGI/EITC/CTC interaction"),
        ("Fiscal weight", "WGTP (household)", "SPM-grain household weight; not the EITC-reweighted tax-unit weight"),
        ("Income quintiles", "WGTP-weighted", "SPM units ranked on summed income; population-equal fifths"),
    ]
    birth_note = (
        (f"Both arms are driven by OBSERVED births — the count of age-0 dependents "
         f"in each PUMS tax unit (an infant <1yr ≈ a birth in the last 12 months, the "
         f"annual flow). The weighted count is calibrated to CDC NVSR resident births "
         f"(by-residence basis, {BIRTH_BASE_YEAR}=15,535), then "
         + (f"projected to {ty} with the repo's damped-trend ensemble "
            f"(point {birth_info['projection']['point']:,.0f}, 90% PI "
            f"[{birth_info['projection']['ci90_low']:,.0f}, "
            f"{birth_info['projection']['ci90_high']:,.0f}])"
            + _birth_nowcast_note(birth_info["projection"]) + "."
            if birth_info.get("trend_applied")
            else "held at the base-year level (--no-birth-projection).")
         + " Each birth draws one prenatal + one postnatal payment. This replaces the "
           "legacy num_dependents × rate proxy, which multiplied an all-dependents "
           "count (incl. older children, students, adult dependents) by a "
           "children-only-calibrated rate and overstated total births ~1.7×.")
        if use_observed else
        ("Both arms are driven by the LEGACY proxy (num_dependents × the annual birth "
         "rate per dependent) — --use-proxy-births. This overstates total births; "
         "prefer the observed-infant basis (default).")
    )
    cc = birth_info.get("county_calibration", {})
    if cc.get("mode") == "doh_county_shares":
        cf = cc["county_factors"]
        county_note = (
            " County split: recalibrated to Hawaiʻi DOH's aggregate county vital-"
            "statistics shares (2018-2025), not raw PUMS county shares — PUMS's own "
            "county allocation is small-sample-noisy (Hawaii County carries only ~11 "
            "sampled infants) and Maui/Kauai are not independently sampled at all "
            "(one combined PUMA, split by a demographic-heuristic imputer, not survey "
            "evidence). Per-county factor vs the raw PUMS-implied count: "
            + "; ".join(f"{c}×{v['factor']:.2f}" for c, v in cf.items())
            + ". The state BIRTH total is unchanged (redistribution only), but "
              "the state COST total is NOT — eligibility is unit-specific and "
              "correlated with county, so this also corrects the eligible share "
              "(measured +2.1% on cost/recipients on the real frame; "
              "eligible_families itself is unchanged). "
              "--no-doh-county-shares restores raw PUMS shares."
        )
    elif cc.get("mode") == "raw_pums_shares":
        county_note = (
            " County split: raw PUMS shares (--no-doh-county-shares) — unreliable for "
            "Hawaii/Maui/Kauai (small samples; Maui/Kauai not independently sampled). "
            "Prefer the default DOH-calibrated split."
        )
    else:
        county_note = ""
    birth_note = birth_note + county_note
    notes = [
        f"RxKids Hawaiʻi — methodology notes (TY {ty})",
        "",
        birth_note + " Cost = the take-up-adjusted benefit weighted by the household "
        "(WGTP) weight, summed to SPM-unit grain. Per-unit amounts are expectations, "
        "not literal payments.",
        "",
        "Eligibility is tested on MAGI (gross income with 100% of Social Security) at "
        "the tax-unit / MAGI-household grain — the family Medicaid defines (42 CFR "
        "435.603) — consistent with the statute's clause-1 Medicaid anchor. If the bill "
        "defines 'family' more broadly, the SPM-family grain (~$45M) is the alternative.",
        "",
        "Household impact here is the distribution of RxKids benefits RECEIVED across "
        "weighted income quintiles (SPM units ranked on summed income). This view does not "
        "use the tax engine or SPM poverty thresholds, so the whole run is genuine TY"
        f"{ty}. A poverty-lift view (persons lifted) would require those engines, which "
        "stop at TY2025 — out of scope for this run.",
        "",
        "Sampling 90% CI is ACS sampling error only (SDR over 80 PUMS replicate weights). "
        "The assumption band is the joint (all-low/all-high corner) sweep over the two "
        "soft drivers — take-up and the birth count (±25%) — cost is linear in each. The "
        "±25% birth band comfortably covers the ensemble birth-projection's ~±23% 90% PI.",
        "",
        "The reported program cost is benefit dollars (cash out the door). The "
        f"administrative line ({100 * args.admin_load:.0f}% of benefit) covers operating "
        "overhead (enrollment, disbursement, eligibility, outreach); the appropriation "
        "total is benefit + admin.",
        "",
        "The base run prices the 6-month postnatal window (Flint's lower bound). The "
        "'potential' line prices the optional extension to the full 12 months — the "
        "program may or may not opt in. Postnatal cost is linear in months, so the "
        "additional 6 months roughly equals the base postnatal arm again.",
        "",
        "Caveats: 2026-2028 FPL is CPI-projected off the 2025 HHS table. The birth "
        "cohort is projected with the ensemble (vital-statistics counts, not survey "
        "estimates); update HI_BIRTHS_BY_YEAR as new NVSR final-data releases land, "
        "and re-pull HI_DOH_BIRTHS_MONTHLY (bump DOH_SNAPSHOT) for the nowcast years. "
        "The occurrence->residence factor is calibrated on the post-2020 travel "
        "regime; a recovery toward the pre-COVID ~1.10 would make the nowcasts high.",
        "",
        "Full methodology: RXKIDS_METHODOLOGY.md (program origin, parameter sourcing, "
        "eligibility approximations, limitations).",
    ]
    ctx = {
        "tax_year": ty,
        "cost_total": total_cost, "cost_prenatal": prenatal_cost,
        "cost_postnatal": postnatal_cost, "cost_se": total_se, "moe": moe,
        "base_postnatal_months": int(p.postnatal_months),
        "ext_postnatal_months": EXTENDED_POSTNATAL_MONTHS,
        "ext_total": ext_total, "ext_postnatal": ext_postnatal,
        "additional_cost": additional_cost,
        "first_year": first_year, "ramp_sensitivity": ramp_sensitivity,
        "eligible_families": eligible_families,
        "rec_total": rec_total, "rec_pregnancies": rec_pregnancies,
        "rec_infants": rec_infants, "avg_benefit": avg_benefit,
        "quintiles": quintiles, "band": band, "county_rows": county_rows,
        "assumptions": assumptions, "notes": notes,
        "admin_load": args.admin_load, "admin_cost": admin_cost,
        "appropriation_total": appropriation_total,
        "birth_info": birth_info,
        "scenario_rows": scenario_rows,
        "county_scenario_rows": county_scenario_rows,
    }

    # ---- Write outputs ----
    workbook_path = args.out / WORKBOOK_NAME
    _write_workbook(workbook_path, ctx=ctx)
    pdf_path = None
    if not args.no_pdf:
        pdf_path = args.out / PDF_NAME
        _write_pdf(pdf_path, ctx=ctx)
    pd.DataFrame(ctx["quintiles"]).to_csv(args.out / "benefits_by_income_quintile.csv", index=False)
    pd.DataFrame(county_rows).to_csv(args.out / "cost_by_county.csv", index=False)
    pd.DataFrame([{
        "geography": "Hawaii", "tax_year": ty,
        "cost_total_$": round(total_cost, 0),
        "cost_prenatal_$": round(prenatal_cost, 0),
        "cost_postnatal_$": round(postnatal_cost, 0),
        "cost_total_se_$": round(total_se, 0),
        "cost_total_ci90_low_$": round(total_cost - moe, 0),
        "cost_total_ci90_high_$": round(total_cost + moe, 0),
        "admin_load_pct": round(100 * args.admin_load, 1),
        "admin_cost_$": round(admin_cost, 0),
        "appropriation_total_$": round(appropriation_total, 0),
        "birth_driver": birth_info["mode"],
        "births_calibrated": round(birth_info["calibrated_weighted"], 0)
        if use_observed else None,
        "births_target": round(birth_info["target"], 0) if use_observed else None,
        "potential_additional_6mo_cost_$": round(additional_cost, 0),
        "potential_total_12mo_cost_$": round(ext_total, 0),
        "first_year_total_$": round(first_year["total"], 0),
        "first_year_prenatal_$": round(first_year["prenatal"], 0),
        "first_year_postnatal_$": round(first_year["postnatal"], 0),
        "first_year_pct_of_steady": round(100 * first_year["pct_of_steady"], 1),
        "eligible_families": round(eligible_families, 0),
        "expected_recipients": round(rec_total, 0),
        "expected_pregnancies": round(rec_pregnancies, 0),
        "expected_infants": round(rec_infants, 0),
        "avg_benefit_per_recipient_$": round(avg_benefit, 0),
    }]).to_csv(args.out / "cost_by_state.csv", index=False)
    # Scenario comparison (state) + county × scenario (tidy long format).
    pd.DataFrame(scenario_rows).to_csv(args.out / "cost_by_scenario.csv", index=False)
    pd.DataFrame(county_scenario_rows).to_csv(
        args.out / "cost_by_county_scenarios.csv", index=False)
    frame.to_parquet(args.out / "spm_units.parquet")

    # ---- Console summary ----
    print(f"\nRxKids Hawaiʻi — TY {ty}")
    print("=" * 56)
    print(f"  Total annual cost          : ${total_cost:>16,.0f}")
    print(f"    Prenatal / Postnatal     : ${prenatal_cost:>14,.0f} / ${postnatal_cost:,.0f}")
    if total_se > 0:
        print(f"  Sampling 90% CI            : ${total_cost - moe:,.0f} – ${total_cost + moe:,.0f}")
    if band is not None:
        print(f"  Assumption band            : ${band['min']:,.0f} – ${band['max']:,.0f}")
    print(f"  Admin load ({100*args.admin_load:.0f}%)          : +${admin_cost:>15,.0f}"
          f"  (appropriation ${appropriation_total:,.0f})")
    if use_observed:
        bi = birth_info
        trend = (f"projected {ty}" if bi.get("trend_applied") else "base level")
        print(f"  Birth driver               : observed age-0 deps, "
              f"{bi['raw_weighted']:,.0f} -> {bi['calibrated_weighted']:,.0f} "
              f"({trend}, ×{bi['factor']:.3f})")
    else:
        print("  Birth driver               : legacy proxy (n_dep × rate)")
    print(f"  POTENTIAL +6 months        : +${additional_cost:>15,.0f}"
          f"  (12-mo total ${ext_total:,.0f})")
    print(f"  FIRST FY (launch)          : ${first_year['total']:>16,.0f}"
          f"  ({100 * first_year['pct_of_steady']:.0f}% of steady; "
          f"{first_year['operating_months']}mo op, {first_year['ramp_months']}mo ramp)")
    print(f"    ramp 6/12/18mo           : ${ramp_sensitivity[6]:,.0f} / "
          f"${ramp_sensitivity[12]:,.0f} / ${ramp_sensitivity[18]:,.0f}")
    print(f"  Eligible families          : {eligible_families:>16,.0f}")
    print(f"  Expected recipients/year   : {rec_total:>16,.0f}"
          f"  ({rec_pregnancies:,.0f} preg + {rec_infants:,.0f} infants)")
    print(f"  Avg benefit per recipient  : ${avg_benefit:>15,.0f}")
    print("\n  Benefit received by income quintile:")
    for row in quintiles:
        print(f"    {row['quintile']}  avg inc ${row['avg_income']:>10,.0f}  "
              f"benefit ${row['total_benefit_$']:>14,.0f}  "
              f"({row['share_of_benefit_pct']:>5.1f}% of total)")
    print("\n  Scenario comparison (benefit cost, take-up held constant):")
    print(f"    {'scenario':<16}{'$/birth':>9}{'cost':>15}{'appropriation':>16}"
          f"{'recipients':>13}")
    for sr in scenario_rows:
        band_s = (f"  band ${sr['assumption_band_low_$']:,.0f}–${sr['assumption_band_high_$']:,.0f}"
                  if sr.get("assumption_band_low_$") is not None else "")
        print(f"    {sr['scenario']:<16}${sr['per_birth_$']:>8,.0f}"
              f"${sr['cost_total_$']:>14,.0f}${sr['appropriation_total_$']:>15,.0f}"
              f"{sr['expected_recipients']:>13,.0f}{band_s}")
    print(f"\n  Workbook: {workbook_path}")
    if pdf_path is not None:
        print(f"  PDF     : {pdf_path}")
    LOG.info("Wrote workbook + %sCSVs + parquet to %s",
             "PDF + " if pdf_path else "", args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
