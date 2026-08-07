"""Tests for the DBEDT MEI fetcher (network-free)."""

from __future__ import annotations

import io
import json
from datetime import datetime

import openpyxl
import pytest

from census_forecaster.markets.screen import (
    HAWAII_PREDICTORS,
    HYPOTHESIS_PAIRS,
    MONTHLY_TARGETS,
)
from census_forecaster.scripts import refresh_dbedt_mei as mei


def _workbook(headers, rows) -> bytes:
    """MEI look-alike: row 3 = series names, col A = timestamps from row 5."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Monthly Economic Indicator")
    ws.cell(row=3, column=1, value="Series")
    for i, h in enumerate(headers):
        ws.cell(row=3, column=2 + i, value=h)
    ws.cell(row=4, column=1, value="UNIT")
    for r, (stamp, vals) in enumerate(rows):
        ws.cell(row=5 + r, column=1, value=stamp)
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(row=5 + r, column=2 + i, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEADERS = ["Civilian labor force 1/", "Visitor arrivals by air 2/",
           "Private Building Permits"]


def test_parses_wanted_series_only():
    out = mei.parse_mei_workbook(_workbook(HEADERS, [
        (datetime(1990, 1, 1), [400000, 555000, 120.5]),
        (datetime(1990, 2, 1), [401000, 500000, 98.0]),
    ]))
    assert set(out) == {"DBEDT_ARRIVALS_", "DBEDT_PERMITS_"}  # not labor force
    assert out["DBEDT_ARRIVALS_"][0] == {"year": 1990, "period": "M01",
                                         "value": 555000.0}
    assert out["DBEDT_PERMITS_"][1]["value"] == 98.0


def test_blank_months_skipped_not_zeroed():
    """Unpublished months (e.g. Maui arrivals after 2026-01) are blank."""
    out = mei.parse_mei_workbook(_workbook(HEADERS, [
        (datetime(2026, 1, 1), [1, 111111, 5.0]),
        (datetime(2026, 2, 1), [1, None, 6.0]),      # arrivals gap
    ]))
    assert len(out["DBEDT_ARRIVALS_"]) == 1
    assert len(out["DBEDT_PERMITS_"]) == 2


def test_non_datetime_rows_ignored():
    out = mei.parse_mei_workbook(_workbook(HEADERS, [
        ("Source: DBEDT", [None, None, None]),
        (datetime(2000, 6, 1), [1, 2.0, 3.0]),
    ]))
    assert len(out["DBEDT_ARRIVALS_"]) == 1


def test_missing_series_raises():
    with pytest.raises(ValueError, match="no wanted series"):
        mei.parse_mei_workbook(_workbook(["Something else"], [
            (datetime(2000, 1, 1), [1.0]),
        ]))


def test_discover_picks_latest_date(monkeypatch):
    html = ('href="https://files.hawaii.gov/dbedt/economic/data_reports/mei/2026-05-honolulu.xlsx"'
            'href="https://files.hawaii.gov/dbedt/economic/data_reports/mei/2026-06-honolulu.xlsx"'
            'href="https://files.hawaii.gov/dbedt/economic/data_reports/mei/2026-06-state.xlsx"')

    class _R:
        text = html
        def raise_for_status(self): pass

    monkeypatch.setattr(mei.requests, "get", lambda *a, **k: _R())
    urls = mei.discover_workbooks()
    assert urls["honolulu"].endswith("2026-06-honolulu.xlsx")   # latest wins
    assert set(urls) == {"honolulu", "state"}


def test_merge_is_additive(tmp_path):
    path = tmp_path / "m.json"
    path.write_text(json.dumps({"series": {"KEEP": [1]}, "sources": {},
                                "limitations": []}))
    merged = mei.merge_into_macro_monthly(
        {"DBEDT_ARRIVALS_STATEWIDE": []}, path=path)
    assert merged["series"]["KEEP"] == [1]
    assert "DBEDT_MEI" in merged["sources"]


def test_screen_repointed_to_current_series():
    """HI_VISITORS must use the DBEDT series (current through today),
    not the HTA workbook series that ends at its final year."""
    assert MONTHLY_TARGETS["HI_VISITORS"][0] == "DBEDT_ARRIVALS_STATEWIDE"
    assert HAWAII_PREDICTORS["HI_VISITORS_ARRIVALS"] == "DBEDT_ARRIVALS_STATEWIDE"
    emitted = {p + g for p in mei.SERIES.values() for g in mei.GEOS.values()}
    assert MONTHLY_TARGETS["HI_VISITORS"][0] in emitted


def test_duplicate_label_resolves_to_distinct_columns():
    """MEI reuses one label for two series: 'Inventory (aver. units on
    market)' appears at BOTH the single-family and condo blocks. A
    first-match-wins lookup silently keeps one and mislabels it; an
    occurrence scan that iterates raw SERIES keys instead of unique
    fragments double-counts columns and collapses both onto the first.
    Both bugs were hit during development — this pins the fix."""
    headers = ["Single-family home resales", "Inventory (aver. units on market)",
               "Condo/Apt/Townhouse units resales",
               "Inventory (aver. units on market)"]
    out = mei.parse_mei_workbook(_workbook(headers, [
        (datetime(2026, 6, 1), [270, 781, 397, 2525]),
    ]))
    sf = out["DBEDT_SF_INVENTORY_"][0]["value"]
    condo = out["DBEDT_CONDO_INVENTORY_"][0]["value"]
    assert (sf, condo) == (781.0, 2525.0), "occurrences collapsed onto one column"


def test_domestic_international_split_parsed():
    """The blended total averages two markets that behave nothing alike.
    Column positions differ between the state book (55/56) and the
    county books (51/52), which is why fragments are matched, not
    indices — so both layouts are exercised here."""
    headers = ["Total visitor days by air", "Domestic visitor days by air",
               "International visitor days by air", "Visitor arrivals by air",
               "Domestic flight visitors 2/", "International flight visitors 2/"]
    out = mei.parse_mei_workbook(_workbook(headers, [
        (datetime(2026, 6, 1), [900, 700, 200, 100, 75, 25]),
    ]))
    assert out["DBEDT_ARRIVALS_"][0]["value"] == 100.0
    assert out["DBEDT_ARRIVALS_DOM_"][0]["value"] == 75.0
    assert out["DBEDT_ARRIVALS_INTL_"][0]["value"] == 25.0
    assert out["DBEDT_VISITOR_DAYS_"][0]["value"] == 900.0
    assert out["DBEDT_VISITOR_DAYS_DOM_"][0]["value"] == 700.0
    assert out["DBEDT_VISITOR_DAYS_INTL_"][0]["value"] == 200.0


def test_total_fragment_does_not_swallow_the_split_columns():
    """'total visitor days by air' must not also match the domestic or
    international variants, and vice versa — substring matching makes
    that a live risk every time a fragment is added."""
    frags = {f for f, _ in mei.SERIES}
    assert "total visitor days by air" in frags
    for label in ("domestic visitor days by air",
                  "international visitor days by air"):
        matched = {f for f in frags if f in label}
        assert matched == {label}, f"{label} also matched {matched - {label}}"


def test_segments_not_screened_against_their_own_total():
    """DBEDT_ARRIVALS_* IS the sum of the DOM and INTL columns (verified
    on the data: worst reconstruction error 1 visitor in 345,075 across
    5 geographies x 438 months). Screening a term against its own sum is
    circular — starker than the HIPHCI case, which at least blended
    four inputs."""
    assert HAWAII_PREDICTORS["HI_VISITORS_INTL"] == "DBEDT_ARRIVALS_INTL_STATEWIDE"
    assert HAWAII_PREDICTORS["HI_VISITORS_DOM"] == "DBEDT_ARRIVALS_DOM_STATEWIDE"
    assert ("HI_VISITORS_INTL", "HI_VISITORS") not in HYPOTHESIS_PAIRS
    assert ("HI_VISITORS_DOM", "HI_VISITORS") not in HYPOTHESIS_PAIRS
    # ...but against labour slack they are legitimate.
    assert ("HI_VISITORS_INTL", "HI_UNEMPLOYMENT") in HYPOTHESIS_PAIRS
    assert ("HI_VISITORS_DOM", "HI_UNEMPLOYMENT") in HYPOTHESIS_PAIRS


def test_no_fragment_is_a_substring_of_another():
    """The general invariant behind two traps found on 2026-08-07:
    'state' matches both "State" payrolls and "State general fund tax
    revenues", and 'agriculture wage and salary jobs' sits inside
    "Total NON-agriculture wage and salary jobs". Substring matching
    makes every new fragment a chance to silently capture a neighbour."""
    frags = sorted({f for f, _ in mei.SERIES})
    collisions = [(a, b) for a in frags for b in frags if a != b and a in b]
    assert collisions == [], f"fragment captured by another: {collisions}"


def test_government_split_and_agriculture_skipped():
    """Both were available and both are traps — pinned so a future
    'completeness' pass does not re-add them without the occurrence
    handling they would need."""
    frags = {f for f, _ in mei.SERIES}
    assert "government" in frags          # the total is safe
    assert "state" not in frags           # collides with tax revenues
    assert "federal" not in frags         # half a split is worse than none
    assert "local" not in frags
    assert not any("agriculture" in f for f in frags)


def test_sector_payrolls_parsed():
    headers = ["Retail Trade", "Professional & Business Services",
               "Health Care & Social Assistance", "Government"]
    out = mei.parse_mei_workbook(_workbook(headers, [
        (datetime(2026, 6, 1), [67000, 82000, 91000, 124000]),
    ]))
    assert out["DBEDT_JOBS_RETAIL_"][0]["value"] == 67000.0
    assert out["DBEDT_JOBS_PROF_"][0]["value"] == 82000.0
    assert out["DBEDT_JOBS_HEALTH_"][0]["value"] == 91000.0
    assert out["DBEDT_JOBS_GOVT_"][0]["value"] == 124000.0


def test_tax_rows_deliberately_not_taken():
    """DOTAX's own reports cover collections at finer granularity with
    revision tracking; two sources for one quantity invites divergence."""
    fragments = {f for f, _ in mei.SERIES}
    for banned in ("general excise", "transient accommodations tax",
                   "state general fund", "wh tax on wages"):
        assert not any(banned in f for f in fragments)
