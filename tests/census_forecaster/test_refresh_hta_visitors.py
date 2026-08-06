"""Tests for the HTA visitor-arrivals fetcher.

Network-free. The load-bearing case is `test_revised_year_headers`:
HTA publishes revised vintages as STRINGS ('2006*', '2010R', '2014R',
'2017R'), so a naive isinstance(int) check silently drops four whole
years — which punches holes through the monthly series and fragments
every Granger window spanning them.
"""

from __future__ import annotations

import io
import json

import openpyxl
import pytest

from census_forecaster.markets.screen import (
    HAWAII_PREDICTORS,
    HYPOTHESIS_PAIRS,
    MONTHLY_TARGETS,
)
from census_forecaster.scripts import refresh_hta_visitors as hta

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _workbook(year_headers, sections=("STATEWIDE", "O'AHU")) -> bytes:
    """Minimal Table 6 look-alike: year headers on a 3-column stride."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hta.SHEET
    ws.cell(row=1, column=1, value="TABLE 6.  VISITOR ARRIVALS BY ISLAND")
    for i, hdr in enumerate(year_headers):
        ws.cell(row=2, column=2 + i * 3, value=hdr)
    row = 4
    for s_i, section in enumerate(sections):
        ws.cell(row=row, column=1, value=section)
        for m_i, month in enumerate(_MONTHS):
            ws.cell(row=row + 1 + m_i, column=1, value=month)
            for y_i in range(len(year_headers)):
                ws.cell(row=row + 1 + m_i, column=2 + y_i * 3,
                        value=1000 + s_i * 100 + m_i)
        row += 13
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parses_sections_and_months():
    out = hta.parse_table6(_workbook([1990, 1991]))
    assert set(out) == {"HTA_VISITORS_STATEWIDE", "HTA_VISITORS_HONOLULU"}
    rows = out["HTA_VISITORS_STATEWIDE"]
    assert len(rows) == 24                       # 2 years x 12 months
    assert rows[0] == {"year": 1990, "period": "M01", "value": 1000.0}
    assert [r["period"] for r in rows[:3]] == ["M01", "M02", "M03"]


def test_revised_year_headers_are_not_dropped():
    """'2006*'/'2010R' are strings — they must still parse as years."""
    out = hta.parse_table6(_workbook([2005, "2006*", "2010R", 2011]))
    years = {r["year"] for r in out["HTA_VISITORS_STATEWIDE"]}
    assert years == {2005, 2006, 2010, 2011}, "revised vintages were dropped"


@pytest.mark.parametrize("value,expected", [
    (1990, 1990), ("2006*", 2006), ("2010R", 2010), (" 2017R ", 2017),
    (2024.0, 2024), ("Total", None), (None, None), ("", None),
    (1800, None), (2200, None), (True, None),
])
def test_header_year(value, expected):
    assert hta._header_year(value) == expected


def test_normalize_handles_okina_and_diacritics():
    assert hta._normalize("O'AHU") == "OAHU"
    assert hta._normalize("KAUA‘I") == "KAUAI"
    assert hta._normalize("LĀNA'I") == "LANAI"
    assert hta._normalize("  HAWAI'I   ISLAND ") == "HAWAII ISLAND"


def test_maui_county_wins_over_maui_island():
    """'MAUI CTY' is the county aggregate; plain 'MAUI' is the island."""
    out = hta.parse_table6(_workbook([1990], sections=("MAUI CTY", "MAUI ")))
    assert "HTA_VISITORS_MAUI" in out
    assert len(out["HTA_VISITORS_MAUI"]) == 12   # only the county section


def test_missing_sheet_raises():
    wb = openpyxl.Workbook()
    wb.active.title = "not-table-6"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="missing"):
        hta.parse_table6(buf.getvalue())


def test_discover_url_finds_historical_workbook(monkeypatch):
    html = ('<a href="/media/1316/2016-highlights.xls">x</a>'
            '<a href="/media/14768/historical-visitors-through-2024-final.xlsx">y</a>')

    class _R:
        text = html
        def raise_for_status(self): pass

    monkeypatch.setattr(hta.requests, "get", lambda *a, **k: _R())
    url = hta.discover_historical_url()
    assert url == hta.HTA_BASE + "/media/14768/historical-visitors-through-2024-final.xlsx"


def test_discover_url_returns_none_when_absent(monkeypatch):
    class _R:
        text = "<a href='/media/1/2016-highlights.xls'>x</a>"
        def raise_for_status(self): pass

    monkeypatch.setattr(hta.requests, "get", lambda *a, **k: _R())
    assert hta.discover_historical_url() is None


def test_merge_is_additive(tmp_path):
    path = tmp_path / "macro_monthly.json"
    existing = {"series": {"LNS14000000": [{"year": 2020, "period": "M01",
                                            "value": 3.5}]},
                "sources": {"BLS": "..."}, "limitations": ["keep me"]}
    path.write_text(json.dumps(existing))
    new = {"HTA_VISITORS_STATEWIDE": [{"year": 2024, "period": "M12",
                                       "value": 700000.0}]}
    merged = hta.merge_into_macro_monthly(new, path=path)
    assert merged["series"]["LNS14000000"] == existing["series"]["LNS14000000"]
    assert merged["series"]["HTA_VISITORS_STATEWIDE"] == new["HTA_VISITORS_STATEWIDE"]
    assert merged["sources"]["BLS"] == "..."
    assert "keep me" in merged["limitations"]


def test_screen_uses_dbedt_but_hta_stays_as_cross_check():
    """The screen's arrivals series moved to DBEDT MEI (current through
    the present month) after the two sources were verified identical to
    within rounding on all 420 overlap months. HTA remains bundled as
    the archival cross-check, so this fetcher's series ids must still be
    emitted — but the screen must NOT point at them anymore."""
    assert MONTHLY_TARGETS["HI_VISITORS"][0] == "DBEDT_ARRIVALS_STATEWIDE"
    assert HAWAII_PREDICTORS["HI_VISITORS_ARRIVALS"] == "DBEDT_ARRIVALS_STATEWIDE"
    emitted = set(hta.SECTIONS.values())
    assert "HTA_VISITORS_STATEWIDE" in emitted
    assert not any(v in emitted for v in HAWAII_PREDICTORS.values())


def test_jets_mechanism_pairs_registered():
    """Both legs of the JETS hypothesis chain are pre-registered."""
    assert ("JETS", "HI_VISITORS") in HYPOTHESIS_PAIRS
    assert ("HI_VISITORS_ARRIVALS", "HI_UNEMPLOYMENT") in HYPOTHESIS_PAIRS
