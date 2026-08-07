"""Refresh DBEDT Monthly Economic Indicator (MEI) county series.

Why this exists
---------------
The HTA historical workbook (``refresh_hta_visitors``) gives clean
monthly visitor arrivals but **ends at the workbook's final year**
(2024). DBEDT's MEI publishes per-county workbooks monthly with a
~5-week lag, each carrying ~52 series **1990-01 → the current month** —
including the same by-air visitor arrivals plus monthly private
building permits, county tax collections, and housing resales. This
fetcher takes the two series with the clearest role today:

* ``DBEDT_ARRIVALS_<GEO>`` — visitor arrivals by air. Extends the
  arrivals signal past HTA's 2024 wall (verified against HTA on the
  1990-2024 overlap before the screen target was repointed).
* ``DBEDT_PERMITS_<GEO>``  — private building permits, monthly. The
  repo's BPS ML channel is annual; this is the same construction signal
  at 12x the cadence, bundled now so a future cadence upgrade has
  history to work with. Not yet screen-wired: "permits → prices" has no
  single clean directional hypothesis (activity vs supply), and the
  pre-registration discipline requires one.

Source & discovery
------------------
``https://dbedt.hawaii.gov/economic/mei/`` links dated workbooks
(``.../data_reports/mei/2026-06-honolulu.xlsx``). There is no stable
"latest" URL, so the listing page is scraped and the newest date wins —
same pattern as ``refresh_hta_visitors``. Keyless.

Workbook layout (single sheet): row 3 = series names, row 4 = units,
col A = month timestamps from row 5.

Usage
-----
    python -m census_forecaster.scripts.refresh_dbedt_mei --dry-run
    python -m census_forecaster.scripts.refresh_dbedt_mei
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Sequence

import requests

from .refresh_zillow_laus_anchors import _atomic_write_json

_PKG_DATA = Path(__file__).resolve().parent.parent / "data"
_MARKETS_DIR = _PKG_DATA / "markets"
MACRO_MONTHLY_FILE = _MARKETS_DIR / "macro_monthly.json"

MEI_PAGE = "https://dbedt.hawaii.gov/economic/mei/"

#: workbook geo slug → series-id suffix (county mapping mirrors
#: refresh_hta_visitors: honolulu=15003, maui=15009, kauai=15007,
#: hawaii=15001).
GEOS: dict[str, str] = {
    "state": "STATEWIDE",
    "honolulu": "HONOLULU",
    "maui": "MAUI",
    "kauai": "KAUAI",
    "hawaii": "HAWAII",
}

#: (header fragment, which occurrence) → emitted series-id prefix.
#:
#: Fragments are matched case-insensitively as substrings of the row-3
#: header. The occurrence index exists because MEI reuses one label for
#: two different series: "Inventory (aver. units on market)" appears at
#: BOTH the single-family block (col 43) and the condo block (col 46).
#: A first-match-wins lookup would silently keep one and mislabel it, so
#: every entry states which occurrence it wants (0-based) and the parser
#: refuses to guess.
SERIES: dict[tuple[str, int], str] = {
    # --- tourism demand (the JETS/arrivals mechanism chain) ---
    ("visitor arrivals by air", 0): "DBEDT_ARRIVALS_",
    ("total visitor days by air", 0): "DBEDT_VISITOR_DAYS_",
    ("visitor expenditures by air", 0): "DBEDT_VISITOR_SPEND_",
    # --- tourism-exposed employment (the labour side of that chain) ---
    ("accommodation", 0): "DBEDT_JOBS_ACCOM_",
    ("food services & drinking places", 0): "DBEDT_JOBS_FOOD_",
    # --- construction activity ---
    ("private building permits", 0): "DBEDT_PERMITS_",
    ("nat. resources, mining, constr", 0): "DBEDT_JOBS_CONSTR_",
    # --- housing transactions (real sales, unlike Zillow's index) ---
    ("single-family home resales", 0): "DBEDT_SF_SALES_",
    ("median selling price", 0): "DBEDT_SF_MEDIAN_",
    ("inventory (aver. units on market)", 0): "DBEDT_SF_INVENTORY_",
    ("condo/apt/townhouse units resales", 0): "DBEDT_CONDO_SALES_",
    ("median price", 0): "DBEDT_CONDO_MEDIAN_",
    ("inventory (aver. units on market)", 1): "DBEDT_CONDO_INVENTORY_",
}

#: Deliberately NOT taken from MEI: the tax-revenue rows (general fund,
#: GE&Use, individual withholding, TAT, county surcharge). DOTAX's own
#: monthly collection reports cover the same ground at finer granularity
#: with explicit revision tracking — see refresh_dotax_collections.py.
#: Two sources for one quantity invites silent divergence.


def discover_workbooks(page_url: str = MEI_PAGE,
                       *, timeout: float = 30.0) -> dict[str, str]:
    """Latest dated workbook URL per geo slug, from the MEI page."""
    resp = requests.get(page_url, timeout=timeout)
    resp.raise_for_status()
    pat = re.compile(
        r'href="(https://files\.hawaii\.gov/dbedt/economic/data_reports/mei/'
        r'(\d{4})-(\d{2})-(state|honolulu|maui|kauai|hawaii)\.xlsx)"',
        re.I,
    )
    best: dict[str, tuple[tuple[int, int], str]] = {}
    for url, yy, mm, geo in pat.findall(resp.text):
        key = geo.lower()
        stamp = (int(yy), int(mm))
        if key not in best or stamp > best[key][0]:
            best[key] = (stamp, url)
    return {geo: url for geo, (_, url) in best.items()}


def parse_mei_workbook(content: bytes) -> dict[str, list[dict]]:
    """One MEI workbook → ``{series_fragment_prefix: rows}``.

    Returns rows keyed by the SERIES prefix (caller appends the geo
    suffix). Values must be numeric; blank cells (months not yet
    published) are skipped.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True,
                                read_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    if len(grid) < 5:
        raise ValueError("workbook too short to be an MEI table")

    headers = grid[2]  # row 3: series names

    # Column index of every occurrence of each wanted fragment, in sheet
    # order — so ("inventory ...", 0) and ("inventory ...", 1) resolve to
    # the single-family and condo columns respectively rather than
    # colliding on the first hit.
    # Scan UNIQUE fragments: two SERIES entries share the "inventory ..."
    # fragment (nth=0 and nth=1), and iterating raw dict keys would append
    # each matching column once per entry — [42, 42, 45, 45] instead of
    # [42, 45] — collapsing both occurrences onto the same column.
    fragments = {fragment for fragment, _ in SERIES}
    occurrences: dict[str, list[int]] = {}
    for idx, name in enumerate(headers):
        label = str(name or "").strip().lower()
        if not label:
            continue
        for fragment in fragments:
            if fragment in label:
                occurrences.setdefault(fragment, []).append(idx)

    col_for_prefix: dict[str, int] = {}
    for (fragment, nth), prefix in SERIES.items():
        cols = occurrences.get(fragment, [])
        if nth < len(cols):
            col_for_prefix[prefix] = cols[nth]
    if not col_for_prefix:
        raise ValueError(
            f"no wanted series found; header sample: "
            f"{[str(h)[:30] for h in headers[:8]]}"
        )

    out: dict[str, list[dict]] = {p: [] for p in col_for_prefix}
    for row in grid[4:]:
        stamp = row[0] if row else None
        if not isinstance(stamp, datetime):
            continue
        for prefix, col in col_for_prefix.items():
            val = row[col] if col < len(row) else None
            if isinstance(val, (int, float)):
                out[prefix].append({
                    "year": stamp.year,
                    "period": f"M{stamp.month:02d}",
                    "value": round(float(val), 2),
                })
    for rows in out.values():
        rows.sort(key=lambda r: (r["year"], r["period"]))
    return {p: r for p, r in out.items() if r}


def merge_into_macro_monthly(series_by_id: dict[str, list[dict]],
                             *, path: Path = MACRO_MONTHLY_FILE) -> dict:
    if path.exists():
        with open(path) as f:
            payload = json.load(f)
    else:
        payload = {"version": 1, "series": {}, "sources": {}, "limitations": []}

    payload.setdefault("series", {}).update(series_by_id)
    payload.setdefault("sources", {})["DBEDT_MEI"] = (
        f"{MEI_PAGE} — Monthly Economic Indicators per-county workbooks"
    )
    payload["fetch_date"] = date.today().isoformat()

    note = (
        "DBEDT_ARRIVALS_* / DBEDT_PERMITS_* come from DBEDT's Monthly "
        "Economic Indicator county workbooks (1990-01 → current month, "
        "~5-week lag, preliminary and revised). ARRIVALS extends the "
        "HTA_VISITORS_* series past the HTA workbook's final year; "
        "PERMITS is monthly county building permits (the BPS ML channel "
        "is annual). Geo suffixes map to FIPS as in HTA_VISITORS_*."
    )
    lims = payload.setdefault("limitations", [])
    if note not in lims:
        lims.append(note)
    return payload


def _parse_args(argv: Optional[Sequence[str]] = None):
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = _parse_args(argv)

    try:
        workbooks = discover_workbooks()
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: could not read the MEI listing page: {exc}",
              file=sys.stderr)
        return 1
    missing = sorted(set(GEOS) - set(workbooks))
    if missing:
        print(f"::warning:: MEI page missing workbooks for: {missing}",
              file=sys.stderr)
    if not workbooks:
        return 1

    series_by_id: dict[str, list[dict]] = {}
    for geo, url in sorted(workbooks.items()):
        suffix = GEOS[geo]
        try:
            resp = requests.get(url, timeout=120)
            resp.raise_for_status()
            parsed = parse_mei_workbook(resp.content)
        except Exception as exc:  # noqa: BLE001 — degrade per-geo
            print(f"::warning:: MEI fetch/parse failed for {geo}: {exc}",
                  file=sys.stderr)
            continue
        for prefix, rows in parsed.items():
            sid = prefix + suffix
            series_by_id[sid] = rows
            print(f"  {sid}: {len(rows)} months "
                  f"({rows[0]['year']}-{rows[0]['period']} → "
                  f"{rows[-1]['year']}-{rows[-1]['period']})", flush=True)

    if not series_by_id:
        print("ERROR: nothing parsed; leaving macro_monthly.json alone.",
              file=sys.stderr)
        return 1

    payload = merge_into_macro_monthly(series_by_id)
    if args.dry_run:
        print(f"[dry-run] would write {len(series_by_id)} series to "
              f"{MACRO_MONTHLY_FILE}")
        return 0
    _atomic_write_json(MACRO_MONTHLY_FILE, payload)
    print(f"Wrote {MACRO_MONTHLY_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
