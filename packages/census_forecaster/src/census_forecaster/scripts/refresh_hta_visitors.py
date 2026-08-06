"""Refresh HTA/DBEDT Hawaii monthly visitor-arrival series.

Why this exists
---------------
The ``JETS`` ticker hypothesis in ``markets/universe.py`` reads: *airline
equity prices embed forward bookings; Hawaii tourism employment follows
visitor arrivals, so JETS should lead the unemployment rate*. The middle
link — actual visitor arrivals — was never in the panel, so the screen
could only test the endpoints and infer the mechanism. This adds the
real monthly arrivals series so the chain can be tested directly.

Source
------
Hawaii Tourism Authority "Historical Visitor Statistics", Table 6
(*Visitor Arrivals by Island and Month*), which carries **1990-2024
monthly** — long enough for the Granger screen's deepest lag (lag-12
needs ~200 aligned months; this supplies ~420). Island rows map onto
Hawaii's four counties:

===================  ==========================  =========
HTA section          County                      FIPS
===================  ==========================  =========
``O'AHU``            Honolulu                    15003
``MAUI CTY``         Maui (incl. Molokaʻi/Lānaʻi) 15009
``KAUA'I``           Kauaʻi                      15007
``HAWAI'I ISLAND``   Hawaiʻi                     15001
===================  ==========================  =========

so these are genuinely **county-level** monthly indicators, unlike the
geoid-constant market/national channels.

Known limitations (see also METHODOLOGY "HTA visitor arrivals")
--------------------------------------------------------------
* **No API.** HTA publishes spreadsheets behind opaque, rotating
  ``/media/<id>/`` URLs, so this script scrapes the listing page to
  discover the current link rather than hardcoding one. UHERO's data
  warehouse does expose a real API but requires a Bearer token, which
  breaks the repo's keyless-fetch discipline.
* **Ends at 2024.** The historical workbook is a yearly snapshot. The
  current-year file is a *separate*, differently-shaped workbook whose
  2026 sheet had real mid-year gaps when checked (Jan/Feb/May present,
  Mar/Apr absent), so it is deliberately NOT merged here — a clean
  1990-2024 backbone beats a ragged edge. Revisit when the next
  "through-<year>" workbook publishes.
* **Revisions.** HTA notes revised 2006/2010/2014/2017 figures; values
  are as-published in the current workbook.

Usage
-----
    python -m census_forecaster.scripts.refresh_hta_visitors --dry-run
    python -m census_forecaster.scripts.refresh_hta_visitors
    python -m census_forecaster.scripts.refresh_hta_visitors --url <xlsx>
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Optional, Sequence

import requests

from .refresh_zillow_laus_anchors import _atomic_write_json

_PKG_DATA = Path(__file__).resolve().parent.parent / "data"
_MARKETS_DIR = _PKG_DATA / "markets"
MACRO_MONTHLY_FILE = _MARKETS_DIR / "macro_monthly.json"

HTA_BASE = "https://www.hawaiitourismauthority.org"
HISTORICAL_PAGE = f"{HTA_BASE}/research/historical-visitor-statistics/"

#: Table 6 sheet name in the historical workbook.
SHEET = "6"

#: HTA section label → emitted series id. Section labels are matched
#: case-insensitively with punctuation stripped, because the workbook
#: uses ʻokina/diacritics inconsistently across vintages.
SECTIONS: dict[str, str] = {
    "STATEWIDE": "HTA_VISITORS_STATEWIDE",
    "OAHU": "HTA_VISITORS_HONOLULU",       # Honolulu County 15003
    "MAUI CTY": "HTA_VISITORS_MAUI",       # Maui County 15009
    "KAUAI": "HTA_VISITORS_KAUAI",         # Kauaʻi County 15007
    "HAWAII ISLAND": "HTA_VISITORS_HAWAII",  # Hawaiʻi County 15001
}

_MONTHS = ("jan", "feb", "mar", "apr", "may", "jun",
           "jul", "aug", "sep", "oct", "nov", "dec")


def _normalize(label: object) -> str:
    """Upper-case, strip diacritics/punctuation/whitespace for matching."""
    text = str(label or "")
    text = (text.replace("‘", "").replace("’", "")
                .replace("ʻ", "").replace("'", ""))
    text = re.sub(r"[ĀÁÀÂ]", "A", text.upper())
    text = re.sub(r"[ĪÍÌÎ]", "I", text)
    text = re.sub(r"[ŌÓÒÔ]", "O", text)
    text = re.sub(r"[ŪÚÙÛ]", "U", text)
    return re.sub(r"\s+", " ", text).strip()


def _header_year(value: object) -> Optional[int]:
    """Year from a Table 6 header cell, or None.

    Accepts ``1990`` (int) and the revised-vintage string forms HTA
    uses — ``'2006*'``, ``'2010R'``, ``'2014R'``, ``'2017R'``.
    """
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        year = int(value)
        return year if 1980 <= year <= 2100 else None
    match = re.match(r"\s*(\d{4})", str(value or ""))
    if not match:
        return None
    year = int(match.group(1))
    return year if 1980 <= year <= 2100 else None


def discover_historical_url(page_url: str = HISTORICAL_PAGE,
                            *, timeout: float = 30.0) -> Optional[str]:
    """Find the 'historical-visitors-through-<year>' workbook link.

    The ``/media/<id>/`` ids rotate on every republish, so the listing
    page is the only stable entry point.
    """
    resp = requests.get(page_url, timeout=timeout)
    resp.raise_for_status()
    hits = re.findall(r'href="(/media/\d+/[^"]*historical-visitors[^"]*\.xlsx)"',
                      resp.text, flags=re.I)
    if not hits:
        return None
    return HTA_BASE + hits[0]


def parse_table6(content: bytes) -> dict[str, list[dict]]:
    """Parse Table 6 → ``{series_id: [{year, period, value}]}``.

    Layout: row 2 holds year headers on a 3-column stride
    (Total / Domestic / International); each section label is followed
    by twelve month rows. Only the Total column is taken.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True,
                                read_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"sheet {SHEET!r} missing; found {wb.sheetnames[:12]}")
    ws = wb[SHEET]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    # Year → column index, from the header row (0-based grid coords).
    #
    # Revised vintages are published as STRINGS with a suffix — '2006*',
    # '2010R', '2014R', '2017R' — so a plain isinstance(int) check
    # silently drops four whole years and punches holes straight through
    # the monthly series (which then fragments every Granger window that
    # spans them). Parse the leading 4-digit year from either form.
    year_col: dict[int, int] = {}
    for row in grid[:6]:
        for idx, val in enumerate(row):
            year = _header_year(val)
            if year is not None:
                year_col.setdefault(year, idx)
        if year_col:
            break
    if not year_col:
        raise ValueError("no year header row found in Table 6")

    wanted = {_normalize(k): sid for k, sid in SECTIONS.items()}
    out: dict[str, list[dict]] = {}

    for r_idx, row in enumerate(grid):
        label = _normalize(row[0] if row else None)
        sid = wanted.get(label)
        if sid is None or sid in out:
            continue                      # first match wins (MAUI CTY vs MAUI)
        rows: list[dict] = []
        for m_off in range(1, 13):
            if r_idx + m_off >= len(grid):
                break
            m_row = grid[r_idx + m_off]
            m_label = _normalize(m_row[0] if m_row else None).lower()[:3]
            if m_label not in _MONTHS:
                break
            month = _MONTHS.index(m_label) + 1
            for year, col in year_col.items():
                if col >= len(m_row):
                    continue
                val = m_row[col]
                if not isinstance(val, (int, float)):
                    continue
                rows.append({"year": int(year), "period": f"M{month:02d}",
                             "value": round(float(val), 2)})
        if rows:
            rows.sort(key=lambda x: (x["year"], x["period"]))
            out[sid] = rows
    return out


def merge_into_macro_monthly(series_by_id: dict[str, list[dict]],
                             *, path: Path = MACRO_MONTHLY_FILE) -> dict:
    """Merge HTA series into macro_monthly.json, preserving the rest."""
    if path.exists():
        with open(path) as f:
            payload = json.load(f)
    else:
        payload = {"version": 1, "series": {}, "sources": {}, "limitations": []}

    payload.setdefault("series", {}).update(series_by_id)
    payload.setdefault("sources", {})["HTA"] = (
        f"{HISTORICAL_PAGE} — Table 6, visitor arrivals by island and month"
    )
    payload["fetch_date"] = date.today().isoformat()

    note = (
        "HTA_VISITORS_* are monthly visitor arrivals (persons) from the HTA "
        "historical workbook, Table 6. County-level: HONOLULU/MAUI/KAUAI/"
        "HAWAII map to FIPS 15003/15009/15007/15001. Ends at the workbook's "
        "final year (2024 as of 2026-08); the current-year HTA file is a "
        "different shape with mid-year gaps and is deliberately not merged. "
        "2006/2010/2014/2017 were revised by HTA."
    )
    lims = payload.setdefault("limitations", [])
    if note not in lims:
        lims.append(note)
    return payload


def _parse_args(argv: Optional[Sequence[str]] = None):
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--url", default=None,
                   help="Workbook URL (default: discover from the HTA page)")
    p.add_argument("--dry-run", action="store_true",
                   help="Fetch and report, but do not write.")
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = _parse_args(argv)

    url = args.url
    if not url:
        try:
            url = discover_historical_url()
        except Exception as exc:  # noqa: BLE001 — degrade, don't crash
            print(f"ERROR: could not read the HTA listing page: {exc}",
                  file=sys.stderr)
            return 1
        if not url:
            print("ERROR: no 'historical-visitors-*.xlsx' link on "
                  f"{HISTORICAL_PAGE} — the page layout may have changed.",
                  file=sys.stderr)
            return 1
    print(f"  workbook: {url}", flush=True)

    try:
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        series = parse_table6(resp.content)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: fetch/parse failed: {exc}", file=sys.stderr)
        return 1

    if not series:
        print("ERROR: parsed zero series; leaving macro_monthly.json alone.",
              file=sys.stderr)
        return 1

    for sid, rows in sorted(series.items()):
        print(f"  {sid}: {len(rows)} months "
              f"({rows[0]['year']}-{rows[0]['period']} → "
              f"{rows[-1]['year']}-{rows[-1]['period']})", flush=True)

    payload = merge_into_macro_monthly(series)
    if args.dry_run:
        print(f"[dry-run] would write {len(series)} series to "
              f"{MACRO_MONTHLY_FILE}")
        return 0

    _atomic_write_json(MACRO_MONTHLY_FILE, payload)
    print(f"Wrote {MACRO_MONTHLY_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
