"""Fetch DOTAX monthly tax-collection reports and merge into bundled disk.

Source: Hawaiʻi Department of Taxation "State Tax Collections and
Distribution" monthly reports, published as XLSX at

    https://files.hawaii.gov/tax/stats/monthly/{YYYYMM}collec.xlsx

plus the fresher GE-specific monthly report ``{YYYYMM}ge.xlsx`` (its
statewide "ALLOCATED COLLECTIONS" row).

Why this exists: these are the closest thing to *revenue ground truth*
the model's targets have — individual withholding is a direct wage-income
nowcast, GE&Use is the broadest activity gauge, TAT tracks tourism.
Everything else in the pipeline is a covariate; this is the outcome.

Publication behavior (observed 2026-07):

* Only roughly the current fiscal year's months stay up as XLSX (older
  months fall back to PDF-only archives), and the full ``collec`` report
  lags ~6-7 months while the GE-specific report lags ~3.
* Every file also carries the same-month-prior-year column, so each
  fetch yields month m *and* m−12.
* Consequently this script MERGES into the existing bundle rather than
  rewriting it: the bundle is the accumulating archive, and later files
  win on overlap (they carry DOTAX's own revisions).

Run it (needs openpyxl — ``pip install census-forecaster[dotax]``):

    python -m census_forecaster.scripts.refresh_dotax_collections
    python -m census_forecaster.scripts.refresh_dotax_collections --months-back 36

Output
------
    src/census_forecaster/data/dotax_monthly/collections.json
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path
from typing import Optional, Sequence

BASE_URL = "https://files.hawaii.gov/tax/stats/monthly"

# collec.xlsx row label (normalized) → series key. Income sub-rows repeat
# identical labels under the "Income-Corp.:" / "Income-Ind.:" section
# headers, so the parser tracks the active section.
_PLAIN_ROWS = {
    "general excise & use": "ge_use",
    "county surcharges": "county_surcharge",
    "trans. accomm. tax/time share occup. tax": "tat",
    "state general fund": "general_fund",
}
_SECTION_ROWS = {
    ("corp", "decl. of est. taxes"): "corp_est",
    ("corp", "payment w/ returns"): "corp_ret",
    ("corp", "refunds"): "corp_refunds",
    ("ind", "decl. of est. taxes"): "ind_est",
    ("ind", "payment w/ returns"): "ind_ret",
    ("ind", "wh tax on wages"): "ind_wh",
    ("ind", "refunds"): "ind_refunds",
}

LIMITATIONS = [
    "Collections are cash-basis by *deposit* month — payment-deadline "
    "timing (estimated-tax due dates, filing season) makes single months "
    "noisy; compare year-over-year same-month or rolling windows only.",
    "Withholding growth conflates wage growth with WITHHOLDING-RATE "
    "policy: the Act 46 (2024) bracket cuts phase in from TY2025 and "
    "mechanically depress withholding relative to wages. GE&Use (rate-"
    "stable) is the cleaner pure-activity gauge.",
    "The full 'collec' report publishes with ~6-7 months of lag and only "
    "the current fiscal year stays up as XLSX; the bundle accumulates "
    "history across refreshes and must never be truncated to one fetch.",
    "ge_allocated comes from the GE-specific report (~3-month lag); it is "
    "'allocated collections', which differs in timing from the collec "
    "report's 'General Excise & Use' row — do not mix the two series.",
    "Later reports revise the prior-year comparison column; on overlap "
    "the value from the most recent report wins (value_asof records the "
    "source file per month).",
]


def _norm(label) -> str:
    """Normalize a row label: strip footnote markers ('1/'), whitespace, case."""
    s = re.sub(r"\s*\d+/\s*$", "", str(label or "").strip())
    return re.sub(r"\s+", " ", s).lower()


def _month_str(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _shift_year(ym: str, delta: int) -> str:
    return f"{int(ym[:4]) + delta}{ym[4:]}"


def _fetch_xlsx(url: str):
    """Return an openpyxl workbook, or None on HTTP 404."""
    import openpyxl
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "census-forecaster/refresh"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)


def parse_collec(wb, year: int, month: int) -> dict[str, dict[str, float]]:
    """Parse one collec.xlsx → {month_str: {series_key: value}} for m and m−12.

    Column layout (verified against Dec-2025): A=label, C=current month,
    D=same month prior year. Cumulative FY columns are ignored — monthly
    values are the primitive; windows are summed downstream.
    """
    ws = wb[wb.sheetnames[0]]
    cur_m = _month_str(year, month)
    prior_m = _shift_year(cur_m, -1)
    out: dict[str, dict[str, float]] = {cur_m: {}, prior_m: {}}
    section = None
    total_seen = False
    for row in ws.iter_rows(min_row=1, max_row=120, max_col=4, values_only=True):
        label = _norm(row[0])
        if not label:
            continue
        if label.startswith("income-corp"):
            section = "corp"
            continue
        if label.startswith("income-ind"):
            section = "ind"
            continue
        key = None
        if (section, label) in _SECTION_ROWS:
            key = _SECTION_ROWS[(section, label)]
        elif label in _PLAIN_ROWS:
            key = _PLAIN_ROWS[label]
            section = None
        elif label.startswith("total -") and not total_seen:
            # First TOTAL row is the current month; its C col is current,
            # and the NEXT total row's D col is prior-year — but prior
            # total also appears on that same next row's C. Simplest
            # robust read: current total from this row, prior from the
            # explicit prior-year row below (label 'total - <prior>').
            key = "total"
            total_seen = True
            if row[2] is not None:
                out[cur_m][key] = round(float(row[2]), 2)
            continue
        elif label.startswith("total -") and total_seen:
            if row[2] is not None:
                out[prior_m]["total"] = round(float(row[2]), 2)
            continue
        if key is None:
            continue
        if row[2] is not None:
            out[cur_m][key] = round(float(row[2]), 2)
        if row[3] is not None:
            out[prior_m][key] = round(float(row[3]), 2)
    return {m: v for m, v in out.items() if v}


def parse_ge(wb, year: int, month: int) -> dict[str, dict[str, float]]:
    """Parse one ge.xlsx → allocated-collections statewide totals for m, m−12."""
    ws = wb[wb.sheetnames[0]]
    cur_m = _month_str(year, month)
    prior_m = _shift_year(cur_m, -1)
    out: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=1, max_row=60, max_col=3, values_only=True):
        if _norm(row[0]) == "allocated collections":
            if row[1] is not None:
                out.setdefault(cur_m, {})["ge_allocated"] = round(float(row[1]), 2)
            if row[2] is not None:
                out.setdefault(prior_m, {})["ge_allocated"] = round(float(row[2]), 2)
            break
    return out


def _default_out() -> Path:
    return (Path(__file__).resolve().parent.parent
            / "data" / "dotax_monthly" / "collections.json")


def _iter_months_back(months_back: int):
    today = date.today()
    y, m = today.year, today.month
    for _ in range(months_back):
        yield y, m
        m -= 1
        if m == 0:
            y, m = y - 1, 12


def refresh(out_path: Path, months_back: int, verbose: bool = True) -> dict:
    existing: dict = {}
    if out_path.exists():
        with open(out_path) as f:
            existing = json.load(f)
    monthly: dict = dict(existing.get("monthly", {}))
    asof: dict = dict(existing.get("value_asof", {}))

    n_files = 0
    # Oldest → newest so that newer files overwrite on overlap.
    for y, m in reversed(list(_iter_months_back(months_back))):
        for kind, parser in (("collec", parse_collec), ("ge", parse_ge)):
            fname = f"{y:04d}{m:02d}{kind}.xlsx"
            wb = _fetch_xlsx(f"{BASE_URL}/{fname}")
            if wb is None:
                continue
            n_files += 1
            if verbose:
                print(f"[dotax] parsed {fname}", file=sys.stderr)
            for month_str, values in parser(wb, y, m).items():
                monthly.setdefault(month_str, {}).update(values)
                asof[month_str] = fname
            wb.close()

    payload = {
        "source": "Hawaii DOTAX monthly collection reports (files.hawaii.gov/tax/stats/monthly)",
        "fetch_date": date.today().isoformat(),
        "units": "USD, cash collections in deposit month",
        "limitations": LIMITATIONS,
        "series_keys": sorted(
            {k for v in monthly.values() for k in v}),
        "value_asof": {k: asof[k] for k in sorted(asof)},
        "monthly": {k: monthly[k] for k in sorted(monthly)},
    }
    return payload, n_files


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Fetch DOTAX monthly collections and merge into the bundle.")
    parser.add_argument("--out", type=Path, default=_default_out())
    parser.add_argument("--months-back", type=int, default=30,
                        help="How many recent report-months to probe (404s skipped).")
    args = parser.parse_args(argv)

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: openpyxl is required — pip install 'census-forecaster[dotax]'",
              file=sys.stderr)
        return 2

    payload, n_files = refresh(args.out, args.months_back)
    if n_files == 0 and not payload["monthly"]:
        print("ERROR: no report files found and no existing bundle — nothing to write.",
              file=sys.stderr)
        return 1

    args.out.parent.mkdir(parents=True, exist_ok=True)
    tmp = args.out.with_suffix(".json.tmp")
    with open(tmp, "w") as f:
        json.dump(payload, f, indent=2, sort_keys=True)
    tmp.replace(args.out)
    print(f"[dotax] {n_files} report files parsed; bundle now spans "
          f"{min(payload['monthly'])} – {max(payload['monthly'])} "
          f"({len(payload['monthly'])} months) → {args.out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
